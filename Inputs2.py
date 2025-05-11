import blpapi
import openpyxl
import shutil
import os
import numpy as np
import time
from datetime import datetime

def setup_bloomberg_session(ticker_symbol):
    """Initialize Bloomberg API session with detailed logging."""
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)
    
    print(f"[INFO] Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("[WARNING] Failed to start Bloomberg session. Ensure Bloomberg Terminal is running.")
        return None
    if not session.openService("//blp/refdata"):
        print("[WARNING] Failed to open Bloomberg reference data service.")
        return None
    print("[INFO] Bloomberg session started successfully.")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=10):
    """Fetch historical data from Bloomberg with timeout and error handling."""
    if len(fields) > 25:
        raise ValueError(f"Too many fields ({len(fields)}). Bloomberg API limit is 25 fields per request.")
    
    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")
    security = f"{ticker} US Equity"
    request.getElement("securities").appendValue(security)
    for field in fields:
        request.getElement("fields").appendValue(field)
    request.set("periodicitySelection", "YEARLY")
    request.set("startDate", f"{start_year}0101")
    request.set("endDate", f"{end_year}1231")
    session.sendRequest(request)
    
    data = {field: {} for field in fields}
    invalid_fields = []
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        event = session.nextEvent(500)
        if event.eventType() in [blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE]:
            for msg in event:
                print(f"[DEBUG] Received message: {msg}")
                if msg.hasElement("responseError"):
                    error = msg.getElement("responseError")
                    error_message = error.getElement("message").getValue()
                    raise ValueError(f"Bloomberg API error: {error_message}")
                if not msg.hasElement("securityData"):
                    raise ValueError("No securityData element in response. Check ticker or data availability.")
                
                security_data = msg.getElement("securityData")
                if security_data.hasElement("fieldExceptions"):
                    field_exceptions = security_data.getElement("fieldExceptions")
                    for i in range(field_exceptions.numValues()):
                        field_error = field_exceptions.getValue(i)
                        invalid_field = field_error.getElement("fieldId").getValue()
                        field_name = field_to_name_map.get(invalid_field, "Unknown Field")
                        invalid_fields.append(invalid_field)
                        print(f"[WARNING] Invalid Bloomberg field detected: {invalid_field} (for '{field_name}')")
                
                field_data = security_data.getElement("fieldData")
                for i in range(field_data.numValues()):
                    datum = field_data.getValue(i)
                    date = datum.getElement("date").getValue()
                    year = date.year
                    for field in fields:
                        if field in invalid_fields:
                            continue
                        if datum.hasElement(field):
                            value = datum.getElement(field).getValue()
                            data[field][year] = value
        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("[WARNING] Bloomberg session terminated unexpectedly.")
                    return None
        elif event.eventType() == blpapi.Event.TIMEOUT:
            print("[DEBUG] Bloomberg event timeout.")
            continue
        if event.eventType() == blpapi.Event.RESPONSE:
            break
    
    if not any(data[field] for field in data):
        print(f"[WARNING] No data received for {ticker} within {timeout}s.")
    if invalid_fields:
        print(f"[INFO] Bloomberg fields skipped due to invalidity: {invalid_fields}")
    return data

def calculate_derived_metrics(data, start_year=2014, end_year=2024):
    """Calculate derived metrics like DSO, DIH, DPO, and Net Working Capital changes."""
    derived = {
        "Changes in Net Working Capital": {},
        "DSO": {},
        "DIH": {},
        "DPO": {},
        "Net Cash from Investments & Acquisitions": {},
        "Increase (Decrease) in Other": {}
    }
    
    for year in range(start_year, end_year + 1):
        if year in data.get("TOT_CUR_ASSETS", {}) and year in data.get("TOT_CUR_LIAB", {}) and \
           year - 1 in data.get("TOT_CUR_ASSETS", {}) and year - 1 in data.get("TOT_CUR_LIAB", {}):
            nwc_t = data["TOT_CUR_ASSETS"][year] - data["TOT_CUR_LIAB"][year]
            nwc_t1 = data["TOT_CUR_ASSETS"][year - 1] - data["TOT_CUR_LIAB"][year - 1]
            derived["Changes in Net Working Capital"][year] = nwc_t - nwc_t1
        
        if year in data.get("BS_ACCOUNTS_RECEIVABLE", {}) and year in data.get("SALES_REV_TURN", {}) and \
           year in data.get("INVENTORIES", {}) and year in data.get("IS_COST_OF_GOODS_SOLD", {}) and \
           year in data.get("ACCT_PAYABLE", {}):
            revenue = data["SALES_REV_TURN"][year]
            cogs = data["IS_COST_OF_GOODS_SOLD"][year]
            ar = data["BS_ACCOUNTS_RECEIVABLE"][year]
            inv = data["INVENTORIES"][year]
            ap = data["ACCT_PAYABLE"][year]
            derived["DSO"][year] = (ar / revenue * 365) if revenue else 0
            derived["DIH"][year] = (inv / cogs * 365) if cogs else 0
            derived["DPO"][year] = (ap / cogs * 365) if cogs else 0
        
        if year in data.get("CF_ACQUISITIONS", {}) and year in data.get("CF_DISPOSALS", {}) and \
           year in data.get("CF_OTHER_INVEST_ACT", {}):
            derived["Net Cash from Investments & Acquisitions"][year] = (
                data["CF_ACQUISITIONS"][year] +
                data["CF_DISPOSALS"][year] +
                data["CF_OTHER_INVEST_ACT"][year]
            )
        
        if year in derived["Changes in Net Working Capital"] and \
           year in data.get("CF_CHG_ACCT_RCV", {}) and year in data.get("CF_CHG_INVENTORIES", {}) and \
           year in data.get("CF_CHG_ACCT_PAYABLE", {}):
            derived["Increase (Decrease) in Other"][year] = (
                derived["Changes in Net Working Capital"][year] -
                (data["CF_CHG_ACCT_RCV"][year] +
                 data["CF_CHG_INVENTORIES"][year] +
                 data["CF_CHG_ACCT_PAYABLE"][year])
            )
    
    return derived

def calculate_cagr(start_value, end_value, years):
    """Calculate Compound Annual Growth Rate."""
    if start_value == 0 or end_value == 0 or years <= 0:
        return 0
    return ((end_value / start_value) ** (1 / years) - 1) * 100

# Updated field map with corrected Bloomberg fields
field_map = {
    # Income Statement (IS)
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "ARDR_DEPRECIATION_AMORTIZATION", "statement": "IS"},
    "Depreciation Expense": {"source": "BDH", "field": "ARDR_DEPRECIATION_EXP", "statement": "IS"},
    "Amortization Expense": {"source": "BDH", "field": "ARDR_AMORT_EXP", "statement": "IS"},
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPERATING_INCOME", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS", "statement": "IS"},
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_BASIC_AVG_SHARES", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_DILUTED_AVG_SHARES", "statement": "IS"},
    # Balance Sheet (BS)
    "Cash & Cash Equivalents & ST Investments": {"source": "BDH", "field": "CASH_AND_ST_INVEST", "statement": "BS"},
    "Cash & Cash Equivalents": {"source": "BDH", "field": "CASH_AND_EQUIV", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCOUNTS_RECEIVABLE", "statement": "BS"},
    "Inventory": {"source": "BDH", "field": "INVENTORIES", "statement": "BS"},
    "Prepaid Expenses and Other Current Assets": {"source": "BDH", "field": "OTH_CUR_ASSETS", "statement": "BS"},
    "Current Assets": {"source": "BDH", "field": "TOT_CUR_ASSETS", "statement": "BS"},
    "Net PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "NET_PPE", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "GROSS_PPE", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "ACCUM_DEPR", "statement": "BS"},
    "Right-of-Use Assets": {"source": "BDH", "field": "OPER_LEASE_ASSETS", "statement": "BS"},
    "Intangibles": {"source": "BDH", "field": "INTANGIBLE_ASSETS", "statement": "BS"},
    "Goodwill": {"source": "BDH", "field": "GOODWILL", "statement": "BS"},
    "Intangibles excl. Goodwill": {"source": "BDH", "field": "NET_OTHER_INTAN_ASSETS", "statement": "BS"},
    "Other Non-Current Assets": {"source": "BDH", "field": "OTH_NON_CUR_ASSETS", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "TOT_NON_CUR_ASSETS", "statement": "BS"},
    "Total Assets": {"source": "BDH", "field": "TOT_ASSETS", "statement": "BS"},
    "Accounts Payable": {"source": "BDH", "field": "ACCT_PAYABLE", "statement": "BS"},
    "Short-Term Debt": {"source": "BDH", "field": "ST_DEBT", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "ST_BORROWINGS", "statement": "BS"},
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "CUR_PORT_LT_LEASE_LIAB", "statement": "BS"},
    "Accrued Expenses and Other Current Liabilities": {"source": "BDH", "field": "OTH_CUR_LIAB", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "TOT_CUR_LIAB", "statement": "BS"},
    "Long-Term Debt": {"source": "BDH", "field": "LT_DEBT", "statement": "BS"},
    # Cash Flow Statement (CF)
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_CHG_ACCT_RCV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHG_INVENTORIES", "statement": "CF", "section": "Operating"},
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMP", "statement": "CF", "section": "Operating"},
    "Other Operating Adjustments": {"source": "BDH", "field": "CF_OTHER_OPER_ADJUSTMENTS", "statement": "CF", "section": "Operating"},
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    "Net Capex": {"source": "BDH", "field": "CF_CAP_EXPEND", "statement": "CF", "section": "Investing"},
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "CF_CAPITAL_EXPEND", "statement": "CF", "section": "Investing"},
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "CF_DISPOSAL_PPE_INTAN", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_ACQUISITIONS", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_DISPOSALS", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_PURCH_LT_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_SALE_LT_INVEST", "statement": "CF", "section": "Investing"},
    "Other Investing Inflows (Outflows)": {"source": "BDH", "field": "CF_OTHER_INVEST_ACT", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    "Lease Payments": {"source": "BDH", "field": "CF_LEASE_PAYMENTS", "statement": "CF", "section": "Financing"},
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_BORROW", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_REPAY_LT_DEBT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_CASH_DIV_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "CF_SHARE_REPURCHASE", "statement": "CF", "section": "Financing"},
    "Other Financing Inflows (Outflows)": {"source": "BDH", "field": "CF_OTHER_FIN_ACT", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_FIN_ACT", "statement": "CF", "section": "Financing"},
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_FX_EFFECT", "statement": "CF", "section": "All"},
    # Additional Fields (BS)
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"},
    "Total Debt": {"source": "BDH", "field": "TOT_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "PREFERRED_EQUITY", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONT_INT", "statement": "BS"},
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"},
    "Total Borrowings": {"source": "BDH", "field": "TOT_BORROWINGS", "statement": "BS"},
    "Total Leases": {"source": "BDH", "field": "TOT_LEASE_LIAB", "statement": "BS"},
    "Net Debt": {"source": "BDH", "field": "NET_DEBT", "statement": "BS"},
    "Effective Tax Rate": {"source": "BDH", "field": "EFF_TAX_RATE", "statement": "BS"},
    # Derived Metrics
    "Changes in Net Working Capital": {"source": "derived", "field": "Changes in Net Working Capital", "statement": "BS"},
    "DSO": {"source": "derived", "field": "DSO", "statement": "IS"},
    "DIH": {"source": "derived", "field": "DIH", "statement": "BS"},
    "DPO": {"source": "derived", "field": "DPO", "statement": "BS"},
    "Net Cash from Investments & Acquisitions": {"source": "derived", "field": "Net Cash from Investments & Acquisitions", "statement": "CF", "section": "Investing"},
    "Increase (Decrease) in Other": {"source": "derived", "field": "Increase (Decrease) in Other", "statement": "CF", "section": "Operating"}
}

# Manual cell mapping for 2014 data
field_cell_map = {
    # Income Statement (IS)
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Other Operating (Income) Expenses": "G11",
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14",
    "Amortization Expense": "G15",
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24",
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",
    # Balance Sheet (BS)
    "Cash & Cash Equivalents & ST Investments": "G32",
    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    "Prepaid Expenses and Other Current Assets": "G37",
    "Current Assets": "G38",
    "Net PP&E (Property, Plant and Equipment)": "G39",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    "Right-of-Use Assets": "G42",
    "Intangibles": "G43",
    "Goodwill": "G44",
    "Intangibles excl. Goodwill": "G45",
    "Other Non-Current Assets": "G46",
    "Non-Current Assets": "G47",
    "Total Assets": "G48",
    "Accounts Payable": "G49",
    "Short-Term Debt": "G50",
    "Short-Term Borrowings": "G50",
    "Current Portion of Lease Liabilities": "G51",
    "Accrued Expenses and Other Current Liabilities": "G52",
    "Current Liabilities": "G53",
    "Long-Term Debt": "G54",
    # Cash Flow Statement (CF)
    "Net Income": "G66",
    "(Increase) Decrease in Accounts Receivable": "G67",
    "(Increase) Decrease in Inventories": "G68",
    "Stock Based Compensation": "G69",
    "Other Operating Adjustments": "G70",
    "Operating Cash Flow": "G71",
    "Increase (Decrease) in Other": "G72",
    "Net Capex": "G73",
    "Acquisition of Fixed & Intangibles": "G74",
    "Disposal of Fixed & Intangibles": "G75",
    "Acquisitions": "G76",
    "Divestitures": "G77",
    "Increase in LT Investment": "G78",
    "Decrease in LT Investment": "G79",
    "Other Investing Inflows (Outflows)": "G80",
    "Investing Cash Flow": "G81",
    "Net Cash from Investments & Acquisitions": "G82",
    "Lease Payments": "G83",
    "Debt Borrowing": "G84",
    "Debt Repayment": "G85",
    "Dividends": "G86",
    "Increase (Repurchase) of Shares": "G87",
    "Other Financing Inflows (Outflows)": "G88",
    "Financing Cash Flow": "G89",
    "Effect of Foreign Exchange": "G90",
    # Additional Fields (BS)
    "Market Capitalization": "G91",
    "Total Debt": "G92",
    "Preferred Stock": "G93",
    "Non-Controlling Interest": "G94",
    "Enterprise Value": "G95",
    "Total Borrowings": "G96",
    "Total Leases": "G97",
    "Net Debt": "G98",
    "Effective Tax Rate": "G99",
    # Other Derived Metrics
    "Changes in Net Working Capital": "G100",
    "DSO": "G101",
    "DIH": "G102",
    "DPO": "G103"
}

def filter_field_map(statement, cf_section=None):
    """Filter field_map to include only fields for the selected statement and CF section."""
    allowed_statements = ["IS", "BS", "CF"]
    if statement not in allowed_statements:
        raise ValueError(f"Invalid statement '{statement}'. Choose IS, BS, or CF.")
    
    if statement == "CF" and cf_section not in [None, "Operating", "Investing", "Financing"]:
        raise ValueError(f"Invalid CF section '{cf_section}'. Choose Operating, Investing, Financing, or None for full CF.")
    
    if statement == "CF":
        if cf_section:
            filtered_map = {k: v for k, v in field_map.items() if v["statement"] == "CF" and (v["section"] == cf_section or v["section"] == "All")}
        else:
            filtered_map = {k: v for k, v in field_map.items() if v["statement"] == "CF"}
    else:
        filtered_map = {k: v for k, v in field_map.items() if v["statement"] == statement}
    
    # Add dependent fields for derived metrics
    derived_fields = {
        "Changes in Net Working Capital": ["TOT_CUR_ASSETS", "TOT_CUR_LIAB"],
        "DSO": ["BS_ACCOUNTS_RECEIVABLE", "SALES_REV_TURN"],
        "DIH": ["INVENTORIES", "IS_COST_OF_GOODS_SOLD"],
        "DPO": ["ACCT_PAYABLE", "IS_COST_OF_GOODS_SOLD"],
        "Net Cash from Investments & Acquisitions": ["CF_ACQUISITIONS", "CF_DISPOSALS", "CF_OTHER_INVEST_ACT"],
        "Increase (Decrease) in Other": ["TOT_CUR_ASSETS", "TOT_CUR_LIAB", "CF_CHG_ACCT_RCV", "CF_CHG_INVENTORIES", "CF_CHG_ACCT_PAYABLE"]
    }
    
    for derived_field, deps in derived_fields.items():
        if derived_field in filtered_map:
            for dep in deps:
                for field, config in field_map.items():
                    if config["field"] == dep and field not in filtered_map:
                        filtered_map[field] = config
    
    bdh_fields = [config["field"] for config in filtered_map.values() if config["source"] == "BDH"]
    if len(bdh_fields) > 25:
        print(f"[WARNING] {statement} ({cf_section or 'Full'}) has {len(bdh_fields)} fields, exceeding Bloomberg's 25-field limit. Trimming to first 25 fields.")
        trimmed_map = {}
        bdh_count = 0
        for field, config in filtered_map.items():
            if config["source"] == "derived" or bdh_count < 25:
                trimmed_map[field] = config
                if config["source"] == "BDH":
                    bdh_count += 1
        filtered_map = trimmed_map
    
    return filtered_map

def get_column_letter(col_index):
    """Convert 1-based column index to letter (e.g., 7 → G)."""
    return openpyxl.utils.get_column_letter(col_index)

def get_next_columns(start_cell, num_columns):
    """Get list of column letters starting from start_cell for num_columns to the right."""
    try:
        col_letter = "".join(c for c in start_cell if c.isalpha())
        row = int("".join(c for c in start_cell if c.isdigit()))
        start_col_index = openpyxl.utils.column_index_from_string(col_letter)
        return [get_column_letter(start_col_index + i) + str(row) for i in range(num_columns)]
    except ValueError:
        raise ValueError(f"Invalid cell reference: {start_cell}")

def populate_valuation_model(template_path, ticker, statement, cf_section=None):
    """Populate the Inputs sheet with Bloomberg data for the given ticker, statement, and CF section."""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file {template_path} not found.")
    
    session = setup_bloomberg_session(ticker)
    if not session:
        return
    
    try:
        # Filter field_map for the selected statement and CF section
        selected_field_map = filter_field_map(statement, cf_section)
        bdh_fields = [v["field"] for k, v in selected_field_map.items() if v["source"] == "BDH"]
        # Create a mapping of Bloomberg fields to human-readable names for error reporting
        field_to_name_map = {v["field"]: k for k, v in selected_field_map.items() if v["source"] == "BDH"}
        
        if statement == "CF" and "Net Income" in selected_field_map and "NET_INCOME" in bdh_fields:
            # Remove NET_INCOME from Bloomberg fetch if IS data is available
            bdh_fields.remove("NET_INCOME")
            selected_field_map.pop("Net Income")
            print("[INFO] Using Net Income from IS (G24) for CF.")
        
        print(f"[INFO] Fetching {len(bdh_fields)} fields for {statement}{f' ({cf_section})' if cf_section else ''}: {bdh_fields}")
        
        bdh_data = fetch_bloomberg_data(session, ticker, bdh_fields, field_to_name_map)
        derived_data = calculate_derived_metrics(bdh_data)
        
        output_path = f"{ticker}_{statement}{f'_{cf_section}' if cf_section else ''}_valuation_model.xlsx"
        shutil.copy(template_path, output_path)
        
        wb = openpyxl.load_workbook(output_path)
        if "Inputs" not in wb.sheetnames:
            raise ValueError("Inputs sheet not found in the template file.")
        ws = wb["Inputs"]
        
        # Map years to columns (2014–2024, 11 years)
        years = list(range(2014, 2025))
        
        # Handle Net Income for CF (copy from IS G24 to G66)
        if statement == "CF" and "Net Income" in field_cell_map:
            start_cell = field_cell_map["Net Income"]
            cells = get_next_columns(start_cell, len(years))
            is_cells = get_next_columns("G24", len(years))
            print(f"[DEBUG] Copying Net Income from IS {is_cells} to CF {cells}")
            for is_cell, cf_cell in zip(is_cells, cells):
                if ws[is_cell].value is not None:
                    ws[cf_cell] = ws[is_cell].value
                    ws[cf_cell].number_format = "#,##0.000"
                else:
                    print(f"[WARNING] No data in {is_cell} for Net Income.")
            # Calculate CAGR for Net Income
            start_value = ws[is_cells[0]].value or 0
            end_value = ws[is_cells[-1]].value or 0
            if start_value and end_value:
                cagr_cell = get_column_letter(openpyxl.utils.column_index_from_string(cells[-1][:-len(str(int(cells[-1][1:])))] + 1)) + cells[0][len(cells[0][:-len(str(int(cells[0][1:])))]):]
                ws[cagr_cell] = calculate_cagr(start_value, end_value, 10) / 100
                ws[cagr_cell].number_format = "0.00%"
                print(f"[DEBUG] CAGR for 'Net Income' written to {cagr_cell}")
        
        # Process remaining fields
        for field, config in selected_field_map.items():
            if field not in field_cell_map:
                print(f"[WARNING] No cell defined for '{field}' in field_cell_map. Skipping.")
                continue
            
            start_cell = field_cell_map[field]
            try:
                cells = get_next_columns(start_cell, len(years))
                print(f"[DEBUG] Writing '{field}' to cells: {cells}")
                
                if config["source"] == "BDH":
                    values = bdh_data.get(config["field"], {})
                    for year, cell in zip(years, cells):
                        if year in values:
                            ws[cell] = values[year] / 1000  # Convert to millions (Bloomberg in thousands)
                            ws[cell].number_format = "#,##0.000"  # Comma as thousand separator, 3 decimals
                    start_value = (values.get(2014, 0) / 1000) if 2014 in values else 0
                    end_value = (values.get(2024, 0) / 1000) if 2024 in values else 0
                elif config["source"] == "derived":
                    values = derived_data[config["field"]]
                    for year, cell in zip(years, cells):
                        if year in values:
                            ws[cell] = values[year]
                            ws[cell].number_format = "#,##0.000"
                    start_value = values.get(2014, 0) or 0
                    end_value = values.get(2024, 0) or 0
                
                # Write CAGR
                if start_value and end_value:
                    cagr_cell = get_column_letter(openpyxl.utils.column_index_from_string(cells[-1][:-len(str(int(cells[-1][1:])))] + 1)) + cells[0][len(cells[0][:-len(str(int(cells[0][1:])))]):]
                    ws[cagr_cell] = calculate_cagr(start_value, end_value, 10) / 100
                    ws[cagr_cell].number_format = "0.00%"
                    print(f"[DEBUG] CAGR for '{field}' written to {cagr_cell}")
            
            except Exception as e:
                print(f"[WARNING] Error writing '{field}' to {start_cell} and subsequent cells: {e}")
        
        wb.save(output_path)
        print(f"[INFO] Valuation model saved as {output_path}")
    
    except Exception as e:
        print(f"[ERROR] Error during Bloomberg API interaction for {ticker}: {e}")
    
    finally:
        try:
            session.stop()
            print("[INFO] Bloomberg session stopped.")
        except Exception as e:
            print(f"[WARNING] Error stopping Bloomberg session: {e}")

if __name__ == "__main__":
    print("[INFO] This script processes one financial statement (IS, BS, or CF) at a time due to Bloomberg's 25-field limit.")
    print("[INFO] For CF, you can select Operating, Investing, or Financing activities, or the full CF statement.")
    print("[INFO] Run the script separately for each statement/section. Close the script after each run.")
    print("[INFO] Output files will be named <ticker>_<statement>[_<section>]_valuation_model.xlsx (e.g., AAPL_CF_Operating_valuation_model.xlsx).")
    
    template_path = "LIS_Valuation_Empty.xlsx"
    
    ticker = input("Enter the ticker symbol (e.g., AAPL): ").strip().upper()
    if not ticker:
        print("[ERROR] Ticker symbol cannot be empty.")
    elif not ticker.isalnum():
        print("[ERROR] Ticker symbol must contain only letters and numbers.")
    else:
        statement = input("Enter financial statement (IS, BS, CF): ").strip().upper()
        cf_section = None
        if statement == "CF":
            cf_section = input("Enter CF section (Operating, Investing, Financing, or press Enter for full CF): ").strip().capitalize()
            if cf_section == "":
                cf_section = None
        try:
            populate_valuation_model(template_path, ticker, statement, cf_section)
        except Exception as e:
            print(f"[ERROR] An error occurred: {e}")
