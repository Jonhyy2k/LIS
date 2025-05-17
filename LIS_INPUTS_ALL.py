# You need to login to the Bloomberg Terminal for the script to work!
# Run it using the arrow on the top right.
# Enter the stock ticker while specifying the country in the end,
# For example AAPL US or 000660 KS (the script will automatically add [Equity]

# This file scecifically was made by GOOGLE GEMINI adapted by the previous code, not checked yet...

import blpapi
import openpyxl
import shutil
import os
import time
from datetime import datetime, timedelta

# --- Bloomberg Session Setup ---
def setup_bloomberg_session(ticker_symbol_for_log=""):
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)

    print(f"\n✨ Attempting to connect to Bloomberg{' for ' + ticker_symbol_for_log if ticker_symbol_for_log else ''}...")
    if not session.start():
        print("⚠️ Whoops! Failed to start the Bloomberg session. Please make sure the Bloomberg Terminal is running and the API is enabled.")
        return None
    if not session.openService("//blp/refdata"):
        print("⚠️ Uh oh! Failed to open the Bloomberg reference data service. Can't get the data without this.")
        session.stop()
        return None
    print(f"✅ Great! Bloomberg session{' for ' + ticker_symbol_for_log if ticker_symbol_for_log else ''} started successfully.")
    return session

# --- Fetch Historical Data (BDH) ---
def fetch_bloomberg_historical_data(session, ticker, fields, field_to_name_map, start_date_str, end_date_str, periodicity="DAILY", overrides=None, timeout=60):
    if not fields:
        print("ℹ️ Just a heads up: No specific historical data fields were requested for this round.")
        return {}

    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")

    security_for_request = f"{ticker} Equity"
    request.getElement("securities").appendValue(security_for_request)

    for field in fields:
        request.getElement("fields").appendValue(field)

    request.set("periodicitySelection", periodicity) # YEARLY, MONTHLY, WEEKLY, DAILY
    request.set("startDate", start_date_str) # YYYYMMDD
    request.set("endDate", end_date_str)     # YYYYMMDD
    request.set("nonTradingDayFillOption", "ALL_CALENDAR_DAYS") # Or ACTIVE_DAYS_ONLY
    request.set("nonTradingDayFillMethod", "PREVIOUS_VALUE")   # Or NIL_VALUE

    # Apply currency override if ticker is non-US for financial data
    # For price data, usually keep local currency unless specified.
    parts = ticker.strip().split()
    country_code = ""
    if len(parts) > 1 and len(parts[-1]) == 2 and parts[-1].isalpha():
        country_code = parts[-1].upper()

    # Only apply USD override if it's explicitly for financial data fields that need it
    # This logic might need refinement based on which fields are being fetched.
    # For PX_LAST, we typically want it in local currency.
    is_financial_data_request = any(f in ["SALES_REV_TURN", "BS_ACCT_NOTE_RCV"] for f in fields) # Example check
    if is_financial_data_request:
        if country_code and country_code != "US":
            print(f"🌍 Looks like a non-US stock ({ticker}, Country: {country_code}) for financial data. I'll ask Bloomberg for data in USD.")
            request.set("currency", "USD")
        elif country_code == "US":
            print(f"🇺🇸 This stock ({ticker}) is US-based for financial data. Data should come in USD by default.")
        else:
            print(f"🤔 Can't quite tell the country for {ticker} for financial data. Requesting in local currency.")


    if overrides:
        for override_field, override_value in overrides.items():
            override = request.getElement("overrides").appendElement()
            override.setElement("fieldId", override_field)
            override.setElement("value", override_value)

    print(f"📡 Sending HistoricalDataRequest to Bloomberg for {security_for_request}. Asking for: {fields} from {start_date_str} to {end_date_str} ({periodicity})")
    session.sendRequest(request)

    data = {field: {} for field in fields} # field: {date: value}
    invalid_fields = []
    start_time = time.time()

    while True:
        if time.time() - start_time >= timeout:
            print(f"⏳ Timeout! It's been {timeout} seconds waiting for historical data for {security_for_request}.")
            # Mark all requested years/dates for non-invalid fields as timeout
            for field_id_timeout in fields:
                if field_id_timeout not in invalid_fields:
                    # For historical, we don't know all dates beforehand easily if it's daily etc.
                    # So, this timeout primarily means the request as a whole failed to complete.
                    # The calling function will need to handle this lack of data.
                    print(f"    Field {field_id_timeout} might be incomplete due to timeout.")
            break

        event = session.nextEvent(500)

        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"🕰️ Bloomberg is taking a moment to respond for historical data ({security_for_request}). Still waiting...")
            continue

        for msg in event:
            if msg.hasElement("responseError"):
                error = msg.getElement("responseError")
                error_message = error.getElement("message").getValue()
                print(f"❌ Error from Bloomberg (HistoricalDataRequest) for {security_for_request}: {error_message}.")
                for f_id_err in fields:
                    if f_id_err not in invalid_fields: invalid_fields.append(f_id_err)
                continue

            if not msg.hasElement("securityData"):
                print(f"🤔 Hmm, historical data message for {security_for_request} missing 'securityData'.")
                continue

            security_data = msg.getElement("securityData")

            if security_data.hasElement("securityError"):
                sec_error = security_data.getElement("securityError")
                error_msg_sec = sec_error.getElement("message").getValueAsString()
                print(f"❌ Security Error (HistoricalDataRequest) for '{security_data.getElementValue('security')}': {error_msg_sec}.")
                for f_id_sec_err in fields:
                    if f_id_sec_err not in invalid_fields: invalid_fields.append(f_id_sec_err)
                continue

            if security_data.hasElement("fieldExceptions"):
                field_exceptions = security_data.getElement("fieldExceptions")
                for j in range(field_exceptions.numValues()):
                    field_error = field_exceptions.getValue(j)
                    invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                    error_info = field_error.getElement("errorInfo")
                    error_message_field = error_info.getElement("message").getValueAsString()
                    field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field")
                    print(f"⚠️ Problem with historical field: '{invalid_field_id}' ('{field_name_display}') for {security_for_request}. Error: {error_message_field}.")
                    if invalid_field_id not in invalid_fields:
                        invalid_fields.append(invalid_field_id)

            if not security_data.hasElement("fieldData"):
                print(f"😕 No 'fieldData' in historical response for {security_for_request}. All fields might be invalid or no data for period.")
                # Mark all fields as N/A for all years if no fieldData at all
                for f_id_no_data in fields:
                    if f_id_no_data not in invalid_fields: # only if not already marked invalid
                        # This part is tricky for daily data, as we don't have a predefined list of dates/years
                        # The calling function will see empty dicts for these fields.
                        pass
                continue

            field_data_array = security_data.getElement("fieldData")
            for k in range(field_data_array.numValues()): # Iterates over dates
                datum = field_data_array.getValue(k)
                if not datum.hasElement("date"):
                    print("🤔 Historical data entry missing 'date'. Skipping.")
                    continue
                
                date_obj = datum.getElement("date").getValueAsDatetime()
                # For yearly data, we use year. For others, use the full date object or format as needed.
                key_date = date_obj.year if periodicity == "YEARLY" else date_obj.strftime('%Y-%m-%d')


                for field_id in fields:
                    if field_id in invalid_fields:
                        data.setdefault(field_id, {})[key_date] = "N/A (Invalid Field)"
                        continue

                    if datum.hasElement(field_id):
                        try:
                            value = datum.getElement(field_id).getValueAsFloat()
                            data.setdefault(field_id, {})[key_date] = value
                        except blpapi.exception.ElementErrorException:
                            try:
                                value_str = datum.getElement(field_id).getValueAsString()
                                data.setdefault(field_id, {})[key_date] = value_str
                                # print(f"📝 Note: Historical field '{field_id}' for {key_date} for {security_for_request} is text: '{value_str}'")
                            except Exception as e_str:
                                print(f"⚠️ Couldn't get historical value for '{field_id}' for {key_date} for {security_for_request} (even as string): {e_str}.")
                                data.setdefault(field_id, {})[key_date] = "N/A (Error reading)"
                        except Exception as e_gen:
                             print(f"⚠️ Error reading historical field '{field_id}' for {key_date} for {security_for_request}: {e_gen}.")
                             data.setdefault(field_id, {})[key_date] = "N/A (Conversion Error)"
                    else:
                        # If field not present for this date, mark as None (will be N/A (Missing) later if not filled)
                        if key_date not in data.get(field_id, {}):
                            data.setdefault(field_id, {})[key_date] = None
            
        if event.eventType() == blpapi.Event.RESPONSE:
            print(f"📬 Received complete historical response for {security_for_request} ({fields}).")
            # Fill N/A for missing dates only if periodicity is yearly, for daily it's too much.
            if periodicity == "YEARLY":
                start_year = int(start_date_str[:4])
                end_year = int(end_date_str[:4])
                for field_id_fill in fields:
                    if field_id_fill not in invalid_fields:
                        for year_fill_val in range(start_year, end_year + 1):
                            if year_fill_val not in data.get(field_id_fill, {}):
                                data.setdefault(field_id_fill, {})[year_fill_val] = "N/A (Missing)"
            break
        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg_status in event: # Renamed msg to msg_status to avoid conflict
                if msg_status.messageType() == blpapi.Name("SessionTerminated"):
                    print("❗CRITICAL: Bloomberg session terminated unexpectedly during historical data fetch!")
                    return None # Critical failure
                print(f"ℹ️ Bloomberg Session/Service Status update: {msg_status.toString().strip()}")
        else:
            print(f"🤖 Unhandled Bloomberg event type: {event.eventType()}. Message: {msg.toString().strip()}")


    if not any(data.get(field) for field in data):
        print(f"💨 No historical data retrieved for any requested field for {ticker} in this batch.")
    if invalid_fields:
        print(f"🚫 For {security_for_request}, these historical fields were problematic: {invalid_fields}")
    return data

# --- Fetch Reference Data (BDP, Bulk) ---
def fetch_bloomberg_reference_data(session, ticker, fields_bdp, field_to_name_map, overrides=None, timeout=30):
    if not fields_bdp:
        print("ℹ️ No specific reference data fields were requested.")
        return {}

    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("ReferenceDataRequest")

    security_for_request = f"{ticker} Equity"
    request.getElement("securities").appendValue(security_for_request)

    for field in fields_bdp:
        request.getElement("fields").appendValue(field)
    
    # Apply overrides if any (e.g., for number of holders)
    if overrides:
        for override_field, override_value in overrides.items():
            override = request.getElement("overrides").appendElement()
            override.setElement("fieldId", override_field)
            override.setElement("value", override_value)
    
    print(f"📡 Sending ReferenceDataRequest to Bloomberg for {security_for_request}. Asking for: {fields_bdp}")
    session.sendRequest(request)

    data = {} # field: value or list of values for bulk data
    start_time = time.time()

    while True:
        if time.time() - start_time >= timeout:
            print(f"⏳ Timeout! It's been {timeout} seconds waiting for reference data for {security_for_request}.")
            for field_id_timeout in fields_bdp:
                if field_id_timeout not in data: # If not already processed
                    data[field_id_timeout] = "N/A (Timeout)"
            break
        
        event = session.nextEvent(500)

        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"🕰️ Bloomberg is taking a moment to respond for reference data ({security_for_request}). Still waiting...")
            continue

        for msg in event:
            if msg.hasElement("responseError"):
                error = msg.getElement("responseError")
                error_message = error.getElement("message").getValue()
                print(f"❌ Error from Bloomberg (ReferenceDataRequest) for {security_for_request}: {error_message}.")
                for f_id_err in fields_bdp: data[f_id_err] = "N/A (Response Error)"
                continue # Should break or return after responseError? For now, marks fields and continues.

            if not msg.hasElement("securityData"): # Should be securityData (plural) for ref data
                print(f"🤔 Hmm, reference data message for {security_for_request} missing 'securityData' array.")
                continue
            
            security_data_array = msg.getElement("securityData") # This is an array
            for i in range(security_data_array.numValues()): # Should only be one for single security request
                security_data = security_data_array.getValue(i)
                
                # Check for security level error first
                if security_data.hasElement("securityError"):
                    sec_error = security_data.getElement("securityError")
                    error_msg_sec = sec_error.getElement("message").getValueAsString()
                    print(f"❌ Security Error (ReferenceDataRequest) for '{security_data.getElementValue('security')}': {error_msg_sec}.")
                    for f_id_sec_err in fields_bdp: data[f_id_sec_err] = "N/A (Security Error)"
                    continue # Next message or break

                # Field exceptions
                if security_data.hasElement("fieldExceptions"):
                    field_exceptions_array = security_data.getElement("fieldExceptions")
                    for j in range(field_exceptions_array.numValues()):
                        field_exception = field_exceptions_array.getValue(j)
                        field_id_error = field_exception.getElement("fieldId").getValueAsString()
                        error_info = field_exception.getElement("errorInfo")
                        message = error_info.getElement("message").getValueAsString()
                        field_name_display = field_to_name_map.get(field_id_error, "Unknown Field")
                        print(f"⚠️ Problem with reference field: '{field_id_error}' ('{field_name_display}') for {security_for_request}. Error: {message}.")
                        data[field_id_error] = "N/A (Field Exception)"

                # Actual field data
                if security_data.hasElement("fieldData"):
                    field_data = security_data.getElement("fieldData")
                    for field_bdp in fields_bdp:
                        if field_bdp in data and data[field_bdp].startswith("N/A"): # Already marked as error
                            continue

                        if field_data.hasElement(field_bdp):
                            element = field_data.getElement(field_bdp)
                            # Handle bulk data (arrays)
                            if element.isArray():
                                values = []
                                for k in range(element.numValues()):
                                    sub_element = element.getValue(k)
                                    # If array elements are complex (have sub-fields)
                                    if sub_element.numElements() > 0:
                                        item_data = {}
                                        for l in range(sub_element.numElements()):
                                            nested_el = sub_element.getElement(l)
                                            item_data[nested_el.name().toString()] = nested_el.getValueAsString() # Or AsFloat etc.
                                        values.append(item_data)
                                    else: # Simple array of strings/numbers
                                        try:
                                            values.append(sub_element.getValueAsFloat())
                                        except:
                                            values.append(sub_element.getValueAsString())
                                data[field_bdp] = values
                            else: # Single value
                                try:
                                    data[field_bdp] = element.getValueAsFloat()
                                except:
                                    try:
                                        data[field_bdp] = element.getValueAsString()
                                    except Exception as e_val:
                                        print(f"⚠️ Could not get value for ref field {field_bdp}: {e_val}")
                                        data[field_bdp] = "N/A (Read Error)"
                        elif field_bdp not in data: # If field not present at all and no error yet
                             data[field_bdp] = "N/A (Not Found)"
                else: # No fieldData element
                    for field_bdp_nf in fields_bdp:
                        if field_bdp_nf not in data : data[field_bdp_nf] = "N/A (No fieldData)"


        if event.eventType() == blpapi.Event.RESPONSE: # End of message stream for this request
            print(f"📬 Received complete reference data response for {security_for_request}.")
            # Fill any remaining unassigned fields as "N/A (Missing in Response)"
            for field_id_fill in fields_bdp:
                if field_id_fill not in data:
                    data[field_id_fill] = "N/A (Missing in Response)"
            break
        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg_status in event:
                if msg_status.messageType() == blpapi.Name("SessionTerminated"):
                    print("❗CRITICAL: Bloomberg session terminated unexpectedly during reference data fetch!")
                    return {f: "N/A (Session Terminated)" for f in fields_bdp} # Critical failure
                print(f"ℹ️ Bloomberg Session/Service Status update: {msg_status.toString().strip()}")
        else:
            print(f"🤖 Unhandled Bloomberg event type during reference data fetch: {event.eventType()}.")

    return data


# --- Calculate Derived Metrics ---
def calculate_derived_metrics(fetched_data, start_year, end_year):
    # This function is for financial statement derived items.
    # Other derived items (like from BDP fields) can be handled directly.
    derived_data = {}
    other_operating_components = [
        "IS_OTHER_OPER_INC",
        "IS_OTHER_OPERATING_EXPENSES" # Assuming this field exists if it's separate
    ]
    derived_data["Total_Other_Operating"] = {}
    for year in range(start_year, end_year + 1):
        total_other = 0
        has_value = False
        missing_component = False
        for field in other_operating_components:
            if field in fetched_data and year in fetched_data[field]:
                value = fetched_data[field].get(year)
                if isinstance(value, (int, float)):
                    total_other += value
                    has_value = True
                elif isinstance(value, str) and "N/A" in value: # If a component is N/A
                    missing_component = True
                    break
            else: # Field or year not present
                missing_component = True
                break
        
        if missing_component:
            derived_data["Total_Other_Operating"][year] = "N/A (Missing Components)"
        elif has_value:
            derived_data["Total_Other_Operating"][year] = total_other
        else: # No values found, but also no explicit N/A components encountered (e.g. fields totally missing)
            derived_data["Total_Other_Operating"][year] = "N/A (No Data for Components)"
    return derived_data

# --- Field Definitions ---
# BDP: Bloomberg Data Point (Current/Reference)
# BDH: Bloomberg Data Historical (Time series)
# BULK: Bloomberg data that returns an array of values/structures
field_map = {
    # --- Summary Sheet Items ---
    "Company Name": {"source": "BDP", "field": "NAME", "sheet": "Summary"},
    "Company Description": {"source": "BDP", "field": "DES_NOTES", "sheet": "Summary"}, # Or BUSINESS_DESCRIPTION
    "SWOT Analysis": {"source": "BDP", "field": "REES_SWOT_ANALYSIS", "sheet": "Summary"}, # Might be empty often
    "Beta": {"source": "BDP", "field": "BETA_ADJ_OVERRIDABLE", "sheet": "Summary"},
    "Market Float %": {"source": "BDP", "field": "EQY_FLOAT_PCT", "sheet": "Summary"},
    "% Held by Insiders": {"source": "BDP", "field": "PCT_INSIDER_SHARES_OUT", "sheet": "Summary"},
    "% Held by Institutions": {"source": "BDP", "field": "PCT_SHARES_OUTST_INSTITUTIONS", "sheet": "Summary"},
    "Short Interest % Float": {"source": "BDP", "field": "SHORT_INT_FLOAT", "sheet": "Summary"},
    "Major Holders Data": {"source": "BULK", "field": "TOP_INSTITUTIONAL_HOLDERS_OWNERSHIP", "sheet": "Summary", "overrides": {"VW_NUMBER_OF_RESULTS": 5}}, # Example for top 5 institutional holders. Name: "Investor_Name", Pct: "Percentage_Held"

    # --- Price Chart Data (Summary Sheet) ---
    "Historical Prices": {"source": "BDH", "field": "PX_LAST", "sheet": "Summary", "periodicity": "DAILY"},

    # --- Segments Sheet Items ---
    "Geographic Segment Names": {"source": "BULK", "field": "GEO_REVENUE_SEGMENT_NAME", "sheet": "Segments"}, # Example, may need adjustment
    "Geographic Segment Revenues": {"source": "BDH_BULK", "field": "GEO_REVENUE", "sheet": "Segments"}, # Array of values per segment, historical
    "Business Segment Names": {"source": "BULK", "field": "PRODUCT_SEGMENT_NAME", "sheet": "Segments"}, # Example
    "Business Segment Revenues": {"source": "BDH_BULK", "field": "PRODUCT_SEGMENT_REVENUE", "sheet": "Segments"}, # Array of values per segment, historical

    # --- Income Statement (IS) - Inputs Sheet ---
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS", "sheet": "Inputs"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS", "sheet": "Inputs"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS", "sheet": "Inputs"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS", "sheet": "Inputs"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS", "sheet": "Inputs"},
    # "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS", "sheet": "Inputs"}, # Replaced by derived
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS", "sheet": "Inputs"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "CF_DEPR_AMORT", "statement": "CF", "sheet": "Inputs"}, # Often taken from CF
    "Depreciation Expense": {"source": "BDH", "field": "IS_DEPRECIATION_EXP", "statement": "IS", "sheet": "Inputs"}, # More specific if available
    "Amortization Expense": {"source": "BDH", "field": "IS_AMORT_EXP", "statement": "IS", "sheet": "Inputs"}, # More specific
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPER_INC", "statement": "IS", "sheet": "Inputs"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS", "sheet": "Inputs"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS", "sheet": "Inputs"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS", "sheet": "Inputs"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS", "sheet": "Inputs"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS", "sheet": "Inputs"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS", "sheet": "Inputs"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS", "sheet": "Inputs"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS", "sheet": "Inputs"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS_CONT_OPS", "statement": "IS", "sheet": "Inputs"}, # BASIC_EPS
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS_CONT_OPS", "statement": "IS", "sheet": "Inputs"},# DILUTED_EPS
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_AVG_NUM_SH_FOR_EPS", "statement": "IS", "sheet": "Inputs"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_DIL_AVG_SHARES_FOR_EPS", "statement": "IS", "sheet": "Inputs"},# IS_SH_FOR_DILUTED_EPS
    "Total Other Operating Components": {"source": "derived", "field": "Total_Other_Operating", "statement": "IS", "sheet": "Inputs"},


    # --- Balance Sheet (BS) - Inputs Sheet ---
    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS", "sheet": "Inputs"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS", "sheet": "Inputs"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS", "sheet": "Inputs"},
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS", "sheet": "Inputs"},
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS", "sheet": "Inputs"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS", "sheet": "Inputs"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS", "sheet": "Inputs"},
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLES", "statement": "BS", "sheet": "Inputs"},
    "Goodwill": {"source": "BDH", "field": "BS_GOODWILL", "statement": "BS", "sheet": "Inputs"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSET", "statement": "BS", "sheet": "Inputs"},
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS", "sheet": "Inputs"},
    "Short-Term Borrowings": {"source": "BDH", "field": "BS_ST_BORROW", "statement": "BS", "sheet": "Inputs"}, # SHORT_TERM_DEBT_DETAILED
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "BS_ST_CAP_LEASE_OBLIG", "statement": "BS", "sheet": "Inputs"}, # ST_CAPITALIZED_LEASE_LIABILITIES
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS", "sheet": "Inputs"},
    "Long-Term Borrowings": {"source": "BDH", "field": "BS_LT_BORROW", "statement": "BS", "sheet": "Inputs"}, # LONG_TERM_BORROWINGS_DETAILED
    "Long-Term Operating Lease Liabilities": {"source": "BDH", "field": "BS_LT_OPER_LEASES", "statement": "BS", "sheet": "Inputs"}, # LT_CAPITALIZED_LEASE_LIABILITIES
    "Non-Current Liabilities": {"source": "BDH", "field": "BS_NON_CUR_LIAB", "statement": "BS", "sheet": "Inputs"}, # NON_CUR_LIAB
    "Non-Controlling Interest": {"source": "BDH", "field": "BS_MINORITY_INTEREST", "statement": "BS", "sheet": "Inputs"}, # MINORITY_NONCONTROLLING_INTEREST
    "Right-of-Use Assets": {"source": "BDH", "field": "BS_OPER_LEASE_RIGHT_OF_USE_ASSET", "statement": "BS", "sheet": "Inputs"}, # TOT_OPER_LEA_RT_OF_USE_ASSETS

    # --- Cash Flow Statement (CF) - Inputs Sheet ---
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCTS_RCV", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # CF_ACCT_RCV_UNBILLED_REV
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_INVENTORIES", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # CF_CHANGE_IN_INVENTORIES
    "(Increase) Decrease in Pre-paid expeses and Other CA": {"source": "BDH", "field": "CF_OTHER_CUR_ASSETS", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # Needs specific field for prepaid
    "Increase (Decrease) in Accounts Payable": {"source": "BDH", "field": "CF_PAYABLES_ACCRUALS", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # CF_CHANGE_IN_ACCOUNTS_PAYABLE
    "Increase (Decrease) in Accrued Revenues and Other CL": {"source": "BDH", "field": "CF_OTHER_CUR_LIABS", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # Needs specific field
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMP", "statement": "CF", "section": "Operating", "sheet": "Inputs"}, # CF_STOCK_BASED_COMPENSATION
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating", "sheet": "Inputs"},
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "CF_CAP_EXPEND", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # ACQUIS_OF_FIXED_INTANG or NET_CAPEX
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "CF_SALE_OF_FIX_ASSET_INTANG", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # DISPOSAL_OF_FIXED_INTANG
    "Acquisitions": {"source": "BDH", "field": "CF_ACQUISITIONS", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # CF_CASH_FOR_ACQUIS_SUBSIDIARIES
    "Divestitures": {"source": "BDH", "field": "CF_DIVESTITURES", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # CF_CASH_FOR_DIVESTITURES
    "Increase in LT Investment": {"source": "BDH", "field": "CF_NET_CHG_INVEST", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # CF_INCR_INVEST (might be purchase/sale separate)
    # "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECR_INVEST", "statement": "CF", "section": "Investing", "sheet": "Inputs"}, # Covered by NET_CHG_INVEST
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing", "sheet": "Inputs"},
    "Lease Payments": {"source": "BDH", "field": "CF_FIN_LEASE_PYMTS", "statement": "CF", "section": "Financing", "sheet": "Inputs"}, # ARDR_REPAYMENT_FINANCE_LEASES
    "Debt Borrowing": {"source": "BDH", "field": "CF_DEBT_ISSUED", "statement": "CF", "section": "Financing", "sheet": "Inputs"}, # CF_LT_DEBT_CAP_LEAS_PROCEEDS
    "Debt Repayment": {"source": "BDH", "field": "CF_DEBT_REPAID", "statement": "CF", "section": "Financing", "sheet": "Inputs"}, # CF_LT_DEBT_CAP_LEAS_PAYMENT
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing", "sheet": "Inputs"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "CF_COMMON_STOCK_ISSUED_REPURCH", "statement": "CF", "section": "Financing", "sheet": "Inputs"}, # PROC_FR_REPURCH_EQTY_DETAILED
    "Financing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_FNC_ACT", "statement": "CF", "section": "Financing", "sheet": "Inputs"}, # CFF_ACTIVITIES_DETAILED
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FX_RATES", "statement": "CF", "section": "All", "sheet": "Inputs"}, # CF_EFFECT_FOREIGN_EXCHANGES
    "Net Changes in Cash": {"source": "BDH", "field": "CF_NET_CHNG_CASH", "statement": "CF", "section": "All", "sheet": "Inputs"},

    # --- Additional Fields (BS) - Inputs Sheet ---
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS", "sheet": "Inputs", "periodicity": "DAILY"}, # Often needed as of a specific date or historical
    "Total Debt": {"source": "BDH", "field": "BS_TOT_DEBT", "statement": "BS", "sheet": "Inputs"}, # SHORT_AND_LONG_TERM_DEBT
    "Preferred Stock": {"source": "BDH", "field": "BS_PFD_STK", "statement": "BS", "sheet": "Inputs"}, # PFD_EQTY_HYBRID_CAPITAL
    "Enterprise Value": {"source": "BDH", "field": "CUR_ENTP_VAL", "statement": "BS", "sheet": "Inputs", "periodicity": "DAILY"}, # ENTERPRISE_VALUE
}

# --- Cell Mappings (update based on your Excel LIS_Valuation_Empty.xlsx structure) ---
field_cell_map = {
    # --- Summary Sheet ---
    "Company Name": "B4", # Summary Sheet
    "Company Description": "B8", # Summary Sheet (merged cell, write to top-left)
    "SWOT Analysis": "B9", # Summary Sheet, placeholder if DES_NOTES is large, or new section. For now, below description.
    "Beta": "B15", # Summary Sheet
    "Market Float %": "B16", # Summary Sheet
    "% Held by Insiders": "B17", # Summary Sheet
    "% Held by Institutions": "B18", # Summary Sheet
    "Short Interest % Float": "B19", # Summary Sheet
    # Major Holders: A23:Holder1Name, B23:Holder1Pct, A24:Holder2Name, B24:Holder2Pct ...
    "Major Holder 1 Name": "A23", "Major Holder 1 Pct": "B23", # Summary Sheet
    "Major Holder 2 Name": "A24", "Major Holder 2 Pct": "B24", # Summary Sheet
    "Major Holder 3 Name": "A25", "Major Holder 3 Pct": "B25", # Summary Sheet
    "Major Holder 4 Name": "A26", "Major Holder 4 Pct": "B26", # Summary Sheet
    "Major Holder 5 Name": "A27", "Major Holder 5 Pct": "B27", # Summary Sheet
    # Historical Prices for Chart: Starts D4 (Date), E4 (Price) on Summary Sheet
    "Historical Price Data Start Cell": "D4", # Summary Sheet (Date column)
    "Historical Price Value Start Cell": "E4", # Summary Sheet (Price column)

    # --- Segments Sheet ---
    # Geographic Segments: Start Row 6 (example). Col A: Type, Col B: Name, Col C-onwards: Yearly Revenue
    "Geo Segment Type Start Cell": "A6", # Segments Sheet
    "Geo Segment Name Start Cell": "B6", # Segments Sheet
    "Geo Segment Revenue Start Cell": "C6", # Segments Sheet (for the first year of data)
    # Business Segments: Below Geographic, e.g., Start Row 15 (example)
    "Bus Segment Type Start Cell": "A15", # Segments Sheet
    "Bus Segment Name Start Cell": "B15", # Segments Sheet
    "Bus Segment Revenue Start Cell": "C15", # Segments Sheet

    # --- Inputs Sheet (Financials) ---
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Total Other Operating Components": "G11", # Derived
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14",
    "Amortization Expense": "G15",
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24",
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",

    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    "Current Assets": "G38",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    "Right-of-Use Assets": "G42",
    "Intangibles": "G43",
    "Goodwill": "G44",
    "Non-Current Assets": "G47",
    "Accounts Payable": "G49",
    "Short-Term Borrowings": "G51",
    "Current Portion of Lease Liabilities": "G52",
    "Current Liabilities": "G54",
    "Long-Term Borrowings": "G56",
    "Long-Term Operating Lease Liabilities": "G57",
    "Non-Current Liabilities": "G59",
    "Non-Controlling Interest": "G62",

    "(Increase) Decrease in Accounts Receivable": "G69",
    "(Increase) Decrease in Inventories": "G70",
    "(Increase) Decrease in Pre-paid expeses and Other CA": "G71", # Placeholder cell
    "Increase (Decrease) in Accounts Payable": "G72",
    "Increase (Decrease) in Accrued Revenues and Other CL": "G73", # Placeholder cell
    "Stock Based Compensation": "G74",
    "Operating Cash Flow": "G76",
    "Acquisition of Fixed & Intangibles": "G78",
    "Disposal of Fixed & Intangibles": "G79",
    "Acquisitions": "G81",
    "Divestitures": "G82",
    "Increase in LT Investment": "G83",
    # "Decrease in LT Investment": "G84", # Covered by net change
    "Investing Cash Flow": "G86",
    "Debt Borrowing": "G87",
    "Debt Repayment": "G88",
    "Lease Payments": "G89",
    "Dividends": "G90",
    "Increase (Repurchase) of Shares": "G91",
    "Financing Cash Flow": "G93",
    "Effect of Foreign Exchange": "G94",
    "Net Changes in Cash": "G95",

    "Market Capitalization": "G99", # This is yearly in this template, but often current. Will fetch latest year.
    "Total Debt": "G101",
    "Preferred Stock": "G102",
    "Enterprise Value": "G104", # Yearly in this template.
}

# --- Helper Functions ---
def batch_fields(fields_to_fetch, batch_size=25):
    unique_fields = sorted(list(set(fields_to_fetch)))
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index):
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    try:
        base_cell_ref_str = str(base_cell_ref) # Ensure it's a string
        col_str = "".join(filter(str.isalpha, base_cell_ref_str))
        row_num = int("".join(filter(str.isdigit, base_cell_ref_str)))
        start_col_idx = openpyxl.utils.column_index_from_string(col_str)
        target_cells = [f"{get_column_letter_from_index(start_col_idx + i)}{row_num}" for i in range(num_years)]
        return target_cells
    except ValueError as e:
        print(f"❌ Error: The cell reference '{base_cell_ref_str}' looks wrong. Could not figure out the column and row: {e}")
        raise
    except Exception as e_cell:
        print(f"❌ Unexpected error parsing cell '{base_cell_ref_str}': {e_cell}")
        raise

# --- Main Population Logic ---
def populate_valuation_model(template_path, output_path, ticker_symbol, current_field_map, current_field_cell_map):
    if not os.path.exists(template_path):
        print(f"❌ Oh no! The Excel template file '{template_path}' wasn't found.")
        raise FileNotFoundError(f"Template file {template_path} not found.")

    try:
        shutil.copy(template_path, output_path)
        print(f"📄 Copied template to '{output_path}'.")
    except Exception as e_copy:
        print(f"❌ Failed to copy template: {e_copy}")
        raise

    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as e_load:
        print(f"❌ Trouble opening the new Excel file '{output_path}': {e_load}")
        raise

    # --- Setup Years for Financial Data ---
    # Financial data years (e.g., for Income Statement, Balance Sheet, Cash Flow on 'Inputs' sheet)
    # The template seems to go up to 2023 in the provided CSV for 'Inputs', with 2014 as the first historical year.
    # Let's assume 10 years of historical data ending in the most recent full year.
    # For example, if running in mid-2024, end_year_financials would be 2023.
    # The template has columns G to P for 10 years. G is 2014, P is 2023.
    current_actual_year = datetime.now().year
    end_year_financials = current_actual_year -1 # Typically, previous full year for financials
    start_year_financials = end_year_financials - 9 # For 10 years of data
    financial_data_years = list(range(start_year_financials, end_year_financials + 1))
    num_financial_data_years = len(financial_data_years)
    
    # For Price Chart Data (5 years daily)
    end_date_prices = datetime.now()
    start_date_prices = end_date_prices - timedelta(days=5*365) # Approx 5 years
    
    # --- Initialize Data Storage ---
    all_fetched_data = {} # Combined store for BDH and BDP/BULK data
                          # For BDH: { "field_code": {year: value} }
                          # For BDP: { "field_code": value }
                          # For BULK: { "field_code": [array_of_values_or_dicts] }

    # --- Create Global Field to Name Map (for error messages) ---
    global_bberg_code_to_excel_name_map = {
        config["field"]: name for name, config in current_field_map.items() if "field" in config
    }
    
    # --- Separate fields by type for fetching ---
    bdh_fields_to_fetch = {} # field_code: {config}
    bdp_fields_to_fetch = {} # field_code: {config}
    bulk_fields_to_fetch = {}# field_code: {config}

    for excel_name, config in current_field_map.items():
        source_type = config.get("source", "").upper()
        bberg_field = config.get("field")
        if not bberg_field: continue

        if source_type == "BDH":
            bdh_fields_to_fetch[bberg_field] = config
        elif source_type == "BDP":
            bdp_fields_to_fetch[bberg_field] = config
        elif source_type == "BULK" or source_type == "BDH_BULK": # BDH_BULK for historical segment data
            bulk_fields_to_fetch[bberg_field] = config
            if source_type == "BDH_BULK": # Also treat as BDH for fetching if needed
                 bdh_fields_to_fetch[bberg_field] = config


    session = None
    try:
        session = setup_bloomberg_session(ticker_symbol)
        if not session:
            raise ConnectionError("Failed to establish Bloomberg session.")

        # --- Fetch BDH Data (Financials and Historical Prices) ---
        print(f"\n🚀 Phase 1a: Fetching Historical Time Series Data (BDH)...")
        unique_bdh_codes = list(bdh_fields_to_fetch.keys())
        # We need to handle different periodicities and date ranges for BDH fields
        # 1. Financials (Yearly)
        yearly_bdh_codes = [code for code, cfg in bdh_fields_to_fetch.items() if cfg.get("periodicity", "YEARLY").upper() == "YEARLY" or cfg.get("statement")]
        if yearly_bdh_codes:
            print(f"   Fetching yearly financial data for {len(yearly_bdh_codes)} fields from {start_year_financials} to {end_year_financials}...")
            fetched_yearly_data = fetch_bloomberg_historical_data(
                session, ticker_symbol, yearly_bdh_codes, global_bberg_code_to_excel_name_map,
                f"{start_year_financials}0101", f"{end_year_financials}1231", "YEARLY"
            )
            if fetched_yearly_data: all_fetched_data.update(fetched_yearly_data)

        # 2. Price Data (Daily for chart)
        daily_bdh_codes_config = {code: cfg for code, cfg in bdh_fields_to_fetch.items() if cfg.get("periodicity", "").upper() == "DAILY"}
        if daily_bdh_codes_config:
            daily_codes = list(daily_bdh_codes_config.keys())
            print(f"   Fetching daily price data for {len(daily_codes)} fields for the last 5 years...")
            fetched_daily_data = fetch_bloomberg_historical_data(
                session, ticker_symbol, daily_codes, global_bberg_code_to_excel_name_map,
                start_date_prices.strftime('%Y%m%d'), end_date_prices.strftime('%Y%m%d'), "DAILY"
            )
            if fetched_daily_data: all_fetched_data.update(fetched_daily_data)
        
        # --- Fetch BDP & BULK Data (Reference Data) ---
        print(f"\n🚀 Phase 1b: Fetching Current/Reference Data (BDP & BULK)...")
        unique_bdp_bulk_codes = list(set(list(bdp_fields_to_fetch.keys()) + list(bulk_fields_to_fetch.keys())))
        
        # Handle overrides for BULK fields (e.g. number of holders)
        bdp_overrides = {}
        for code, cfg in bulk_fields_to_fetch.items():
            if "overrides" in cfg:
                # For ReferenceDataRequest, overrides are set directly on the request, not per field.
                # This assumes all overrides are compatible. If field-specific overrides are needed,
                # multiple requests might be required.
                # For now, let's assume overrides in field_map are for the general request.
                bdp_overrides.update(cfg["overrides"])


        if unique_bdp_bulk_codes:
            fetched_reference_data = fetch_bloomberg_reference_data(
                session, ticker_symbol, unique_bdp_bulk_codes, global_bberg_code_to_excel_name_map, overrides=bdp_overrides
            )
            if fetched_reference_data: all_fetched_data.update(fetched_reference_data)

    except Exception as e_fetch_main:
        print(f"❌ An error occurred during data fetching phases: {e_fetch_main}")
    finally:
        if session:
            try:
                session.stop()
                print("🔌 Bloomberg session stopped.")
            except Exception as e_stop:
                print(f"⚠️ Minor issue stopping Bloomberg session: {e_stop}")
    
    print(f"\n🏁 Data Fetching Complete. Proceeding to calculations and Excel writing.")

    # --- Calculate Derived Metrics (Financials) ---
    print(f"\n🧮 Phase 2: Calculating derived financial metrics...")
    # Ensure only financial data (which is yearly and in all_fetched_data) is passed
    financial_bdh_data_for_calc = {
        field: data
        for field, data in all_fetched_data.items()
        if field_map.get(global_bberg_code_to_excel_name_map.get(field), {}).get("statement") and isinstance(data, dict)
    }

    derived_financial_data = calculate_derived_metrics(financial_bdh_data_for_calc, start_year_financials, end_year_financials)
    if derived_financial_data:
        all_fetched_data.update(derived_financial_data) # Add to the main data pool
    print("✅ Derived financial metrics calculated.")

    # --- Write Data to Excel ---
    print(f"\n✍️ Phase 3: Writing all data to Excel workbook...")

    # Helper to write a single value
    def write_value_to_cell(sheet_obj, cell_ref, value, number_format=None):
        try:
            if cell_ref: # Ensure cell_ref is valid
                # Convert common N/A representations
                if isinstance(value, str) and value.startswith("N/A"):
                    pass # Keep as string "N/A..."
                elif value is None:
                    value = "N/A (Missing)"
                
                sheet_obj[cell_ref] = value
                if number_format and not isinstance(value, str): # Apply format if not a string (like "N/A")
                    sheet_obj[cell_ref].number_format = number_format
                elif isinstance(value, (int, float)): # Default number format if not specified
                     sheet_obj[cell_ref].number_format = "#,##0.00" if abs(value) >= 1000 else "0.00"


        except Exception as e_write:
            print(f"⚠️ Problem writing value '{str(value)[:50]}' to cell {cell_ref} on sheet '{sheet_obj.title}': {e_write}")


    # --- Populate Summary Sheet ---
    if "Summary" in wb.sheetnames:
        ws_summary = wb["Summary"]
        print("   Populating 'Summary' Sheet...")
        ws_summary[field_cell_map["Company Name"]] = all_fetched_data.get(field_map["Company Name"]["field"], "N/A")
        ws_summary[field_cell_map["Company Description"]] = all_fetched_data.get(field_map["Company Description"]["field"], "N/A")
        
        # SWOT - might be long, ensure cell can handle it or it's truncated.
        swot_text = all_fetched_data.get(field_map["SWOT Analysis"]["field"], "SWOT Analysis not available or field not found.")
        write_value_to_cell(ws_summary, field_cell_map.get("SWOT Analysis"), swot_text)

        write_value_to_cell(ws_summary, field_cell_map["Beta"], all_fetched_data.get(field_map["Beta"]["field"]), "0.00")
        write_value_to_cell(ws_summary, field_cell_map["Market Float %"], all_fetched_data.get(field_map["Market Float %"]["field"]), "0.00'%'")
        write_value_to_cell(ws_summary, field_cell_map["% Held by Insiders"], all_fetched_data.get(field_map["% Held by Insiders"]["field"]), "0.00'%'")
        write_value_to_cell(ws_summary, field_cell_map["% Held by Institutions"], all_fetched_data.get(field_map["% Held by Institutions"]["field"]), "0.00'%'")
        write_value_to_cell(ws_summary, field_cell_map["Short Interest % Float"], all_fetched_data.get(field_map["Short Interest % Float"]["field"]), "0.00'%'")

        # Major Holders
        holders_data = all_fetched_data.get(field_map["Major Holders Data"]["field"])
        if isinstance(holders_data, list):
            for i, holder in enumerate(holders_data[:5]): # Max 5 holders
                if isinstance(holder, dict):
                    # Field names within the 'holder' dict depend on the exact BULK field used.
                    # Common ones are 'Investor_Name' or 'Holder Name', and 'Percentage_Held' or 'Percent Outstanding'
                    # Adjust these keys based on what Bloomberg returns for TOP_INSTITUTIONAL_HOLDERS_OWNERSHIP
                    holder_name = holder.get("Name", holder.get("Investor_Name", f"Holder {i+1} Name N/A"))
                    holder_pct = holder.get("% Out", holder.get("Percentage_Held", "N/A"))
                    
                    name_cell = field_cell_map.get(f"Major Holder {i+1} Name")
                    pct_cell = field_cell_map.get(f"Major Holder {i+1} Pct")
                    if name_cell: ws_summary[name_cell] = holder_name
                    if pct_cell: write_value_to_cell(ws_summary, pct_cell, holder_pct, "0.00'%'") # Assuming pct is a number
        else:
            print(f"   ⚠️ Major Holders data not found or not in expected list format: {holders_data}")
            for i in range(5):
                 name_cell = field_cell_map.get(f"Major Holder {i+1} Name")
                 pct_cell = field_cell_map.get(f"Major Holder {i+1} Pct")
                 if name_cell: ws_summary[name_cell] = "N/A"
                 if pct_cell: ws_summary[pct_cell] = "N/A"


        # Historical Prices for Chart
        price_data = all_fetched_data.get(field_map["Historical Prices"]["field"])
        if isinstance(price_data, dict) and price_data:
            date_cell_start_ref = field_cell_map["Historical Price Data Start Cell"]
            price_cell_start_ref = field_cell_map["Historical Price Value Start Cell"]
            start_col_date_str = "".join(filter(str.isalpha, date_cell_start_ref))
            start_row_date = int("".join(filter(str.isdigit, date_cell_start_ref)))
            start_col_price_str = "".join(filter(str.isalpha, price_cell_start_ref))
            
            current_row = start_row_date
            # Sort dates for chronological order in chart data
            sorted_dates = sorted(price_data.keys(), key=lambda d: datetime.strptime(d, '%Y-%m-%d'))

            ws_summary[f"{start_col_date_str}{current_row-1}"] = "Date" # Header
            ws_summary[f"{start_col_price_str}{current_row-1}"] = "Price" # Header

            for date_str in sorted_dates:
                price_val = price_data[date_str]
                ws_summary[f"{start_col_date_str}{current_row}"] = datetime.strptime(date_str, '%Y-%m-%d').strftime('%Y-%m-%d') # Store as text to preserve format or Excel date
                ws_summary[f"{start_col_date_str}{current_row}"].number_format = 'yyyy-mm-dd'
                write_value_to_cell(ws_summary, f"{start_col_price_str}{current_row}", price_val, "#,##0.00")
                current_row += 1
            print(f"   ✅ Historical price data for chart written to Summary sheet. Please create chart manually from range {date_cell_start_ref}:{price_cell_start_ref.replace(str(start_row_date), '')}{current_row-1}.")
        else:
            print(f"   ⚠️ Historical price data for chart not found or empty.")
            ws_summary[field_cell_map["Historical Price Data Start Cell"]] = "Price Data N/A"

    else:
        print("⚠️ 'Summary' sheet not found in workbook.")

    # --- Populate Inputs Sheet (Financials) ---
    if "Inputs" in wb.sheetnames:
        ws_inputs = wb["Inputs"]
        print("   Populating 'Inputs' Sheet (Financials)...")
        for excel_name, config in current_field_map.items():
            if config.get("sheet") != "Inputs": continue # Only process fields for the Inputs sheet

            base_cell_ref = current_field_cell_map.get(excel_name)
            if not base_cell_ref:
                print(f"   🤔 No Excel cell mapping for '{excel_name}' on Inputs sheet. Skipping.")
                continue

            try:
                # For financials, years are 2014-2023 as per template (10 years)
                # The financial_data_years list (e.g. 2014-2023) must match Excel column order
                target_cells_for_item = get_target_cells_for_years(base_cell_ref, num_financial_data_years)
            except Exception as e_cell_calc:
                print(f"   ❌ Error calculating target cells for '{excel_name}': {e_cell_calc}. Skipping.")
                continue

            bberg_field_code = config.get("field")
            if not bberg_field_code:
                print(f"   🤔 Item '{excel_name}' has no Bloomberg field code. Skipping.")
                continue
            
            data_for_item_yearly = all_fetched_data.get(bberg_field_code, {})

            for i, year_to_populate in enumerate(financial_data_years): # financial_data_years is e.g. [2014, 2015 ... 2023]
                cell_to_write = target_cells_for_item[i]
                raw_value = data_for_item_yearly.get(year_to_populate) # Get data for that specific year

                number_fmt = "#,##0.000"
                if "EPS" in excel_name or "Rate" in excel_name: number_fmt = "0.00"
                if "Shares" in excel_name: number_fmt = "#,##0"
                
                # Handle Market Cap and EV which might be fetched daily but template expects yearly
                if excel_name in ["Market Capitalization", "Enterprise Value"] and isinstance(raw_value, dict):
                    # If it's daily data, try to get value for year-end or latest available in that year
                    # This is a simplification; ideally, you'd specify the exact date for historical MktCap/EV.
                    # For now, if it's a dict (daily data), mark as "See Daily" or try to find a value.
                    # For simplicity, if the BDH request for these was yearly, it would be fine.
                    # If it was daily, this logic needs to be smarter or the request changed.
                    # Assuming for now these were fetched yearly as per template structure.
                    pass # Value should be direct if fetched yearly.

                write_value_to_cell(ws_inputs, cell_to_write, raw_value, number_fmt)
    else:
        print("⚠️ 'Inputs' sheet not found in workbook.")


    # --- Populate Segments Sheet ---
    # This is complex due to the array nature and matching to yearly columns.
    # The Bloomberg fields GEO_REVENUE / PRODUCT_SEGMENT_REVENUE often return
    # an array of structures, where each structure has segment name and then
    # an array of values for different fiscal periods.
    # Example structure from GEO_REVENUE:
    # [ {"Segment Name": "USA", "Fiscal Year": [2023, 2022, ...], "Revenue": [1000, 900, ...]}, ... ]
    # This requires careful parsing.
    # For now, I'll add placeholders and a note that this part needs more specific implementation
    # based on the exact structure returned by your Bloomberg fields.
    if "Segments" in wb.sheetnames:
        ws_segments = wb["Segments"]
        print("   Populating 'Segments' Sheet (Placeholder - requires specific parsing)...")
        # TODO: Implement detailed segment data population.
        # This involves:
        # 1. Fetching fields like `GEO_SEGMENT_DATA` or `PRODUCT_SEGMENT_DATA` (these are hypothetical, actual fields vary).
        #    These fields often return bulk data with segment names and corresponding revenue arrays for multiple years.
        # 2. Parsing this bulk data.
        # 3. Matching the years from Bloomberg with the year columns in your Excel sheet.
        # 4. Writing "Geographic" or "Business" in "Segment Type" column.
        # 5. Writing segment names.
        # 6. Writing revenues into the correct year columns.
        
        # Example of how one might start, assuming `all_fetched_data` contains parsed segment data:
        # geo_segment_data = all_fetched_data.get(field_map["Geographic Segment Revenues"]["field"], [])
        # bus_segment_data = all_fetched_data.get(field_map["Business Segment Revenues"]["field"], [])
        #
        # current_row_segments = int("".join(filter(str.isdigit, field_cell_map["Geo Segment Type Start Cell"])))
        #
        # For Geographic Segments:
        # for segment in geo_segment_data: # Assuming segment is a dict {"Segment Name": "X", "RevenuesByYear": {2023: Y, 2022: Z}}
        #     ws_segments[f"A{current_row_segments}"] = "Geographic"
        #     ws_segments[f"B{current_row_segments}"] = segment.get("Segment Name", "N/A")
        #     # ... then loop through financial_data_years to populate C, D, E... cells
        #     current_row_segments +=1
        #
        # (Similar logic for Business Segments)

        ws_segments["A4"] = "Segment Data Population: Manual implementation needed based on Bloomberg field structure."
        ws_segments["A5"] = "See comments in script for guidance on fetching and parsing segment data."
        print("   ⚠️ Segment data population is highly dependent on specific Bloomberg fields and their output structure. Placeholder added.")

    else:
        print("⚠️ 'Segments' sheet not found in workbook.")


    # --- Save Workbook ---
    try:
        wb.save(output_path)
        print(f"\n🎉 All Done! Your valuation model has been populated and saved to: '{output_path}'")
    except Exception as e_save:
        print(f"❌ Critical Error: Failed to save the final Excel workbook to '{output_path}'. Error: {e_save}")
        print("   Possible reasons: The file might be open in Excel, or there might be a permissions issue.")


if __name__ == "__main__":
    print("-" * 70)
    print(" ✨ Bloomberg Data to Excel Valuation Model Populator (Extended) ✨ ")
    print("-" * 70)

    excel_template_path = "LIS_Valuation_Empty.xlsx"

    try:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_folder_name = "Bloomberg_Valuation_Reports_Extended"
        output_directory = os.path.join(desktop_path, output_folder_name)
        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"📂 Created new output directory: '{output_directory}'")
        else:
            print(f"📂 Using existing output directory: '{output_directory}'")
    except Exception as e_path_create:
        print(f"⚠️ Could not create/access Desktop output directory: {e_path_create}. Saving to script directory.")
        output_directory = "."

    ticker_input = ""
    while not ticker_input:
        raw_input_str = input("➡️ Enter Ticker Symbol (e.g., AAPL US or 000660 KS): ").strip()
        if raw_input_str and any(char.isalnum() for char in raw_input_str):
            ticker_input = raw_input_str.upper()
        else:
            print("❗ Please enter a valid ticker symbol.")

    safe_ticker_filename = ticker_input.replace(" ", "_").replace("/", "_")
    output_file_name = f"{safe_ticker_filename}_Valuation_Model_Ext_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    final_output_path = os.path.join(output_directory, output_file_name)

    print(f"\n📝 Using Template: '{excel_template_path}'")
    print(f"💾 Final Report Will Be Saved As: '{final_output_path}'")
    print(f"🎯 Ticker for Bloomberg: '{ticker_input}' (Script will append ' Equity')")

    try:
        print("\n⏳ Starting data population process... This may take some time.\n")
        populate_valuation_model(
            template_path=excel_template_path,
            output_path=final_output_path,
            ticker_symbol=ticker_input,
            current_field_map=field_map,
            current_field_cell_map=field_cell_map
        )
    except FileNotFoundError as e_fnf:
        print(f"❌ CRITICAL ERROR: File not found. {e_fnf}")
    except ConnectionError as e_conn:
        print(f"❌ CRITICAL ERROR: Bloomberg connection issue. {e_conn}")
        print("   Ensure Bloomberg Terminal is running, you are logged in, and API is configured (bbcomm).")
    except blpapi.exception.BlpException as e_blp:
        print(f"❌ CRITICAL BLPAPI ERROR: {e_blp}")
    except Exception as e_main_exc:
        print(f"❌ AN UNEXPECTED CRITICAL ERROR occurred: {e_main_exc}")
        import traceback
        print("\n--- Technical Error Details ---")
        traceback.print_exc()
        print("--- End Technical Error Details ---\n")
    finally:
        print("\n👋 Script execution finished.")
        print("   Review the output file. Some data, especially Segments, may need manual verification or adjustments based on specific Bloomberg field availability.")
        print("   For the stock chart, data has been placed in the 'Summary' sheet; please create the chart in Excel using that data.")

