import os
import sys
import json
import numpy as np
import pandas as pd
import yfinance as yf
import requests
from datetime import datetime
from scipy.stats import linregress
import matplotlib.pyplot as plt
import openpyxl
import shutil
from openpyxl.utils import get_column_letter, column_index_from_string


class DCFModel:
    def __init__(self, ticker):
        """
        Initialize the DCF model with the company ticker
        """
        self.ticker = ticker.upper()
        self.today = datetime.now().strftime('%Y-%m-%d')
        self.data = {}
        self.projections = {} 
        self.dcf_inputs = {}
        self.dcf_results = {} 
        self.company_info = {}
        self.historical_years = [] # List of pandas Timestamps, e.g., [Timestamp('2020-12-31'), ...]
        self.projection_years = [] # List of 'YYYY' strings, e.g., ['2023', '2024', ...]
        self.historical = {} # Dict with 'YYYY' string keys, e.g., {'2020': {...}, ...}
        self.wacc = None

        self.income_stmt = pd.DataFrame()
        self.balance_sheet = pd.DataFrame()
        self.cash_flow = pd.DataFrame()
        self.price_history = pd.DataFrame()
        self.beta = 1.0

        self.dcf_params = {
            'projection_years': 5, 'terminal_growth_rate': 0.02, 'risk_free_rate': 0.04,
            'market_risk_premium': 0.055, 'capex_to_revenue': 0.02, 
            'working_capital_to_revenue_change': 0.26, 'acquisition_to_revenue': 0.02,
            'dividend_payout': 0.30, 'share_repurchases': 0.25,
        }

    def download_data(self):
        print(f"Retrieving data for {self.ticker}...")
        try:
            self.stock = yf.Ticker(self.ticker)
            self.company_info = {
                'name': self.stock.info.get('shortName', self.ticker),
                'sector': self.stock.info.get('sector', 'Unknown'),
                'industry': self.stock.info.get('industry', 'Unknown'),
                'description': self.stock.info.get('longBusinessSummary', 'No description available'),
                'country': self.stock.info.get('country', 'Unknown'),
                'employees': self.stock.info.get('fullTimeEmployees', 0),
                'market_cap': self.stock.info.get('marketCap'), 
                'shares_outstanding': self.stock.info.get('sharesOutstanding', 0),
                'current_price': self.stock.info.get('currentPrice', self.stock.info.get('previousClose', 0)),
                'currency': self.stock.info.get('currency', 'USD')
            }
            self.income_stmt = self.stock.income_stmt.fillna(0) if not self.stock.income_stmt.empty else pd.DataFrame()
            self.balance_sheet = self.stock.balance_sheet.fillna(0) if not self.stock.balance_sheet.empty else pd.DataFrame()
            self.cash_flow = self.stock.cashflow.fillna(0) if not self.stock.cashflow.empty else pd.DataFrame()
            self.price_history = self.stock.history(period="5y")
            beta_val = self.stock.info.get('beta') 
            self.beta = float(beta_val) if beta_val is not None else 1.0

            if not self.income_stmt.empty:
                 self.historical_years = sorted(self.income_stmt.columns.tolist())[-3:] 
            else: self.historical_years = []
            print(f"Data retrieved. Historical years (timestamps): {[yr.strftime('%Y-%m-%d') for yr in self.historical_years]}")
        except Exception as e:
            print(f"Error downloading data for {self.ticker}: {str(e)}")

    def calculate_historical_metrics(self):
        print(f"Calculating historical metrics for {self.ticker}...")
        if not self.historical_years:
            print("Warning: No historical years. Cannot calculate metrics."); self.historical = {}; return

        self.historical = {ts.strftime('%Y'): {} for ts in self.historical_years}
        for year_ts in self.historical_years: 
            year_str = year_ts.strftime('%Y') 
            
            revenue = 0.0
            if not self.income_stmt.empty and year_ts in self.income_stmt.columns:
                revenue_val = self.income_stmt.get('Total Revenue', pd.Series(dtype=float)).get(year_ts, 0.0)
                if pd.isna(revenue_val) or revenue_val == 0.0: revenue_val = self.income_stmt.get('Revenue', pd.Series(dtype=float)).get(year_ts, 0.0)
                revenue = float(revenue_val) if not pd.isna(revenue_val) else 0.0
            self.historical[year_str]['revenue'] = revenue

            ebit = 0.0
            if not self.income_stmt.empty and year_ts in self.income_stmt.columns:
                ebit_val = self.income_stmt.get('Operating Income', pd.Series(dtype=float)).get(year_ts, 0.0)
                if pd.isna(ebit_val) or ebit_val == 0.0: ebit_val = self.income_stmt.get('EBIT', pd.Series(dtype=float)).get(year_ts, 0.0)
                ebit = float(ebit_val) if not pd.isna(ebit_val) else 0.0
            self.historical[year_str]['ebit'] = ebit
            
            net_income_val = 0.0
            if not self.income_stmt.empty and 'Net Income' in self.income_stmt.index and year_ts in self.income_stmt.columns:
                net_income_val_df = self.income_stmt.loc['Net Income', year_ts]
                net_income_val = float(net_income_val_df) if not pd.isna(net_income_val_df) else 0.0
            self.historical[year_str]['net_income'] = net_income_val
            
            tax_prov = 0.0; pretax_inc_final = ebit 
            if not self.income_stmt.empty and year_ts in self.income_stmt.columns:
                tax_prov_val = self.income_stmt.get('Tax Provision', pd.Series(dtype=float)).get(year_ts, 0.0)
                tax_prov = float(tax_prov_val) if not pd.isna(tax_prov_val) else 0.0
                
                pretax_inc_stmt_val = self.income_stmt.get('Pretax Income', pd.Series(dtype=float)).get(year_ts) 
                if pretax_inc_stmt_val is not None and not pd.isna(pretax_inc_stmt_val) : 
                    pretax_inc_final = float(pretax_inc_stmt_val) if float(pretax_inc_stmt_val) != 0 else (ebit if ebit !=0 else 0.00001) # Avoid div by zero if all zero
                elif ebit != 0 : 
                     pretax_inc_final = ebit
                elif ebit == 0 and pretax_inc_final == 0: 
                     pretax_inc_final = 1 # Avoid division by zero if all components are zero


            tax_rate_calc = abs(tax_prov / pretax_inc_final) if pretax_inc_final and not pd.isna(pretax_inc_final) and pretax_inc_final != 0 and not pd.isna(tax_prov) else 0.25
            self.historical[year_str]['tax_rate'] = tax_rate_calc if 0 <= tax_rate_calc <= 1 else 0.25

            capex_val = 0.0
            if not self.cash_flow.empty and 'Capital Expenditure' in self.cash_flow.index and year_ts in self.cash_flow.columns:
                capex_val_df = self.cash_flow.loc['Capital Expenditure', year_ts]
                capex_val = float(capex_val_df) if not pd.isna(capex_val_df) else 0.0
            self.historical[year_str]['capex'] = abs(capex_val)
            
            dep = 0.0
            if not self.cash_flow.empty and 'Depreciation' in self.cash_flow.index and year_ts in self.cash_flow.columns:
                dep_cf = self.cash_flow.loc['Depreciation', year_ts]
                if not pd.isna(dep_cf): dep = float(dep_cf)
            if (pd.isna(dep) or dep == 0.0) and not self.income_stmt.empty and 'Depreciation And Amortization' in self.income_stmt.index and year_ts in self.income_stmt.columns:
                dep_is = self.income_stmt.loc['Depreciation And Amortization', year_ts]
                if not pd.isna(dep_is): dep = float(dep_is)
            self.historical[year_str]['depreciation'] = dep

            curr_assets = 0.0; curr_liab = 0.0
            if not self.balance_sheet.empty and year_ts in self.balance_sheet.columns:
                curr_assets_val = self.balance_sheet.get('Total Current Assets', pd.Series(dtype=float)).get(year_ts, 0.0)
                curr_liab_val = self.balance_sheet.get('Total Current Liabilities', pd.Series(dtype=float)).get(year_ts, 0.0)
                curr_assets = float(curr_assets_val) if not pd.isna(curr_assets_val) else 0.0
                curr_liab = float(curr_liab_val) if not pd.isna(curr_liab_val) else 0.0
            self.historical[year_str]['working_capital'] = curr_assets - curr_liab

            lt_debt = 0.0; st_debt = 0.0
            if not self.balance_sheet.empty and year_ts in self.balance_sheet.columns:
                lt_debt_val = self.balance_sheet.get('Long Term Debt', pd.Series(dtype=float)).get(year_ts, 0.0)
                st_debt_val = self.balance_sheet.get('Short Term Debt', pd.Series(dtype=float)).get(year_ts, 0.0)
                lt_debt = float(lt_debt_val) if not pd.isna(lt_debt_val) else 0.0
                st_debt = float(st_debt_val) if not pd.isna(st_debt_val) else 0.0
            self.historical[year_str]['total_debt'] = lt_debt + st_debt
            
            cash_val = 0.0
            if not self.balance_sheet.empty and 'Cash And Cash Equivalents' in self.balance_sheet.index and year_ts in self.balance_sheet.columns:
                 cash_val_df = self.balance_sheet.loc['Cash And Cash Equivalents', year_ts]
                 cash_val = float(cash_val_df) if not pd.isna(cash_val_df) else 0.0
            self.historical[year_str]['cash'] = cash_val
            self.historical[year_str]['ebit_margin'] = self.historical[year_str]['ebit'] / self.historical[year_str]['revenue'] if self.historical[year_str]['revenue'] else 0.0

        hist_strs_sorted = [ts.strftime('%Y') for ts in self.historical_years]
        for i in range(1, len(hist_strs_sorted)):
            curr_yr, prev_yr = hist_strs_sorted[i], hist_strs_sorted[i-1]
            if prev_yr in self.historical and self.historical[prev_yr].get('revenue'): 
                self.historical[curr_yr]['revenue_growth'] = (self.historical[curr_yr]['revenue'] / self.historical[prev_yr]['revenue']) - 1
            else: self.historical[curr_yr]['revenue_growth'] = 0.0
            
            wc_change = self.historical[curr_yr]['working_capital'] - self.historical[prev_yr]['working_capital']
            rev_change = self.historical[curr_yr]['revenue'] - self.historical[prev_yr]['revenue']
            self.historical[curr_yr]['wc_to_revenue_change'] = wc_change / rev_change if rev_change else 0.0
        print("Historical metrics calculated.")

    def generate_projections(self):
        print(f"Generating projections for {self.ticker}...")
        if not self.historical or not self.historical_years: print("Error: Hist. data missing."); return
        
        hist_strs = sorted(list(self.historical.keys()))
        if not hist_strs: print("Error: No historical year strings after processing."); return 
        latest_hist_yr_str = hist_strs[-1]
        
        rev_growths = [self.historical[yr].get('revenue_growth', 0.03) for yr in hist_strs[1:] if 'revenue_growth' in self.historical[yr]]
        avg_rev_growth = np.mean(rev_growths) if rev_growths and not pd.isna(np.mean(rev_growths)) else 0.03

        ebit_margins_list = [self.historical[yr]['ebit_margin'] for yr in hist_strs if 'ebit_margin' in self.historical[yr]]
        avg_ebit_margin = np.mean(ebit_margins_list) if ebit_margins_list and not pd.isna(np.mean(ebit_margins_list)) else 0.10
        
        tax_rates_list = [self.historical[yr]['tax_rate'] for yr in hist_strs if 'tax_rate' in self.historical[yr] and 0 <= self.historical[yr]['tax_rate'] <= 1]
        avg_tax_rate = np.mean(tax_rates_list) if tax_rates_list and not pd.isna(np.mean(tax_rates_list)) else 0.25
        
        dep_rev_ratios = [self.historical[yr]['depreciation'] / self.historical[yr]['revenue'] for yr in hist_strs if self.historical[yr].get('revenue') and 'depreciation' in self.historical[yr]]
        avg_dep_to_rev = np.mean(dep_rev_ratios) if dep_rev_ratios and not pd.isna(np.mean(dep_rev_ratios)) else 0.03
        
        capex_rev_ratios = [self.historical[yr]['capex'] / self.historical[yr]['revenue'] for yr in hist_strs if self.historical[yr].get('revenue') and 'capex' in self.historical[yr]]
        avg_capex_to_rev = np.mean(capex_rev_ratios) if capex_rev_ratios and not pd.isna(np.mean(capex_rev_ratios)) else self.dcf_params['capex_to_revenue']

        wc_rev_chg_ratios = [self.historical[yr]['wc_to_revenue_change'] for yr in hist_strs[1:] if 'wc_to_revenue_change' in self.historical[yr]]
        avg_wc_to_rev_chg = np.mean(wc_rev_chg_ratios) if wc_rev_chg_ratios and not pd.isna(np.mean(wc_rev_chg_ratios)) else self.dcf_params['working_capital_to_revenue_change']

        current_year = int(latest_hist_yr_str)
        self.projection_years = [str(current_year + i + 1) for i in range(self.dcf_params['projection_years'])]
        self.projections = {year: {} for year in self.projection_years}
        latest_revenue = self.historical[latest_hist_yr_str]['revenue']

        for i, year_str in enumerate(self.projection_years):
            factor = (self.dcf_params['projection_years'] - i) / self.dcf_params['projection_years']
            growth_rate = avg_rev_growth * factor + self.dcf_params['terminal_growth_rate'] * (1 - factor)
            prev_revenue = self.projections[self.projection_years[i-1]]['revenue'] if i > 0 else latest_revenue
            
            self.projections[year_str]['revenue'] = prev_revenue * (1 + growth_rate)
            self.projections[year_str]['ebit'] = self.projections[year_str]['revenue'] * avg_ebit_margin
            self.projections[year_str]['taxes'] = self.projections[year_str]['ebit'] * avg_tax_rate
            self.projections[year_str]['nopat'] = self.projections[year_str]['ebit'] - self.projections[year_str]['taxes']
            self.projections[year_str]['depreciation'] = self.projections[year_str]['revenue'] * avg_dep_to_rev
            self.projections[year_str]['capex'] = self.projections[year_str]['revenue'] * avg_capex_to_rev
            rev_change = self.projections[year_str]['revenue'] - prev_revenue
            self.projections[year_str]['working_capital_change'] = rev_change * avg_wc_to_rev_chg
            self.projections[year_str]['fcf'] = self.projections[year_str]['nopat'] + self.projections[year_str]['depreciation'] - \
                                               self.projections[year_str]['capex'] - self.projections[year_str]['working_capital_change']
        print("Projections generated.")

    def calculate_wacc(self):
        print(f"Calculating WACC for {self.ticker}...")
        market_cap_val = self.company_info.get('market_cap') 
        if not self.historical or market_cap_val is None or pd.isna(market_cap_val) or not self.historical_years: 
             self.wacc = 0.10; print(f"WACC defaulted to {self.wacc:.2%}"); return self.wacc

        latest_hist_yr_str = sorted(list(self.historical.keys()))[-1]
        cost_of_equity = self.dcf_params['risk_free_rate'] + self.beta * self.dcf_params['market_risk_premium']
        market_cap = float(market_cap_val) 
        total_debt = self.historical[latest_hist_yr_str].get('total_debt',0.0)
        total_capital = market_cap + total_debt

        if total_capital == 0.0: self.wacc = cost_of_equity; print(f"WACC (CoE as Total Capital is zero): {self.wacc:.2%}"); return self.wacc

        weight_equity = market_cap / total_capital; weight_debt = total_debt / total_capital
        cost_of_debt_pre_tax = self.dcf_params['risk_free_rate'] + 0.03 
        tax_rate = self.historical[latest_hist_yr_str].get('tax_rate',0.25)
        cost_of_debt_after_tax = cost_of_debt_pre_tax * (1 - tax_rate)
        self.wacc = (weight_equity * cost_of_equity) + (weight_debt * cost_of_debt_after_tax)
        print(f"WACC calculated: {self.wacc:.2%}"); return self.wacc

    def calculate_dcf(self):
        print(f"Calculating DCF for {self.ticker}...")
        if self.wacc is None or not self.projections or not self.projection_years:
            print("Error: WACC/Projections missing."); self.dcf_results = {}; return {}

        pv_fcf_sum = sum(self.projections[yr]['fcf']/((1+self.wacc)**(i+1)) for i,yr in enumerate(self.projection_years))
        last_fcf = self.projections[self.projection_years[-1]]['fcf']
        safe_tgr = min(self.dcf_params['terminal_growth_rate'], self.wacc - 0.001) 
        tv_denominator = self.wacc - safe_tgr
        if tv_denominator <= 0 : 
            print(f"Warning: WACC ({self.wacc:.2%}) is <= safe_tgr ({safe_tgr:.2%}). TV might be unrealistic.")
            tv = last_fcf * (1 + safe_tgr) / 0.001 
        else: tv = (last_fcf * (1 + safe_tgr)) / tv_denominator
            
        pv_tv = tv / ((1 + self.wacc) ** len(self.projection_years))
        ev = pv_fcf_sum + pv_tv
        
        latest_hist = sorted(list(self.historical.keys()))[-1]
        net_debt = self.historical[latest_hist].get('total_debt',0.0) - self.historical[latest_hist].get('cash',0.0)
        equity_val = ev - net_debt
        shares_val = self.company_info.get('shares_outstanding') 
        shares = float(shares_val) if shares_val is not None and shares_val > 0 else 0.0
        per_share_val = equity_val / shares if shares else 0.0 
            
        self.dcf_results = {'enterprise_value':ev, 'equity_value':equity_val, 'per_share_value':per_share_val,
                            'current_price': self.company_info.get('current_price',0), 'sum_pv_fcf': pv_fcf_sum, 'pv_terminal_value': pv_tv}
        print(f"DCF calculated. EV: {ev:,.0f}, Share Value: {per_share_val:.2f}"); return self.dcf_results

    def run_full_analysis(self, output_dir="./"):
        try:
            self.download_data()
            self.calculate_historical_metrics()
            self.generate_projections()
            self.calculate_wacc()
            self.calculate_dcf()
            print(f"Full analysis complete for {self.ticker}.")
            return True
        except Exception as e:
            print(f"Full analysis error for {self.ticker}: {str(e)}"); import traceback; traceback.print_exc(); return False

    def populate_excel_template(self, template_path, output_path):
        print(f"Populating Excel for {self.ticker} from {template_path} to {output_path}")
        
        # Prerequisite checks
        required_attrs_non_empty_list_dict = ['projection_years', 'historical', 'projections', 
                                              'company_info', 'dcf_results', 'ticker']
        for attr in required_attrs_non_empty_list_dict:
            val = getattr(self, attr, None)
            if val is None: 
                print(f"Error: Attribute '{attr}' is None. Run analysis first."); return
            if isinstance(val, (list, dict)) and not val and attr != 'historical_years': 
                print(f"Error: Attribute '{attr}' is empty. Run analysis first."); return
        if self.wacc is None: print("Error: WACC not calculated."); return
        
        if not hasattr(self, 'historical_years'): 
             print(f"Error: Attribute 'historical_years' not found. Run download_data first."); return

        for stmt_name in ['income_stmt', 'balance_sheet', 'cash_flow']: 
            if not hasattr(self, stmt_name): 
                 print(f"Error: Financial statement DataFrame '{stmt_name}' not found. Run download_data."); return

        try:
            shutil.copy(template_path, output_path)
            wb = openpyxl.load_workbook(output_path)
            inputs_sheet = wb["Inputs"] if "Inputs" in wb.sheetnames else None
            if not inputs_sheet: print("Error: 'Inputs' sheet missing."); wb.close(); return
            dcf_sheet = wb["DCF"] if "DCF" in wb.sheetnames else None
            if dcf_sheet: dcf_sheet['R4'] = self.ticker
            else: print("Warning: 'DCF' sheet missing.")

            field_cell_map = {
                 "Revenue (Sales)": "G6", "COGS (Cost of Goods Sold)": "G7", "Gross Profit": "G8",
                 "SG&A (Selling, General & Administrative)": "G9", "R&D (Research & Development)": "G10",
                 "Total Other Operating Components": "G11", "EBITDA": "G12", "D&A (Depreciation & Amortization)": "G13",
                 "Depreciation Expense": "G14", "Amortization Expense": "G15", "Operating Income (EBIT)": "G16",
                 "Net Interest Expense (Income)": "G17", "Interest Expense": "G18", "Interest Income": "G19",
                 "FX (Gain) Loss": "G20", "Other Non-Operating (Income) Expenses": "G21", "Pre-Tax Income (EBT)": "G22",
                 "Tax Expense (Benefits)": "G23", "Net Income": "G24", "EPS Basic": "G25", "EPS Diluted": "G26",
                 "Basic Weighted Average Shares": "G27", "Diluted Weighted Average Shares": "G28",
                 "Cash & Cash Equivalents": "G33", "Short-Term Investments": "G34", "Accounts Receivable": "G35",
                 "Inventory": "G36", "Current Assets": "G38", "Gross PP&E (Property, Plant and Equipment)": "G40",
                 "Accumulated Depreciation": "G41", "Right-of-Use Assets": "G42", "Intangibles": "G43", "Goodwill": "G44",
                 "Non-Current Assets": "G47", "Accounts Payable": "G49", "Short-Term Borrowings": "G51",
                 "Current Portion of Lease Liabilities": "G52", "Current Liabilities": "G54", "Long-Term Borrowings": "G56",
                 "Long-Term Operating Lease Liabilities": "G57", "Non-Current Liabilities": "G59", "Non-Controlling Interest": "G62",
                 "(Increase) Decrease in Accounts Receivable": "G69", "(Increase) Decrease in Inventories": "G70",
                 "(Increase) Decrease in Pre-paid expeses and Other CA": "G71", "Increase (Decrease) in Accounts Payable": "G72",
                 "Increase (Decrease) in Accrued Revenues and Other CL": "G73", "Stock Based Compensation": "G74",
                 "Operating Cash Flow": "G76", "Acquisition of Fixed & Intangibles": "G78", "Disposal of Fixed & Intangibles": "G79",
                 "Acquisitions": "G81", "Divestitures": "G82", "Increase in LT Investment": "G83", "Decrease in LT Investment": "G84",
                 "Investing Cash Flow": "G86", "Debt Borrowing": "G87", "Debt Repayment": "G88", "Lease Payments": "G89",
                 "Dividends": "G90", "Increase (Repurchase) of Shares": "G91", "Financing Cash Flow": "G93",
                 "Effect of Foreign Exchange": "G94", "Market Capitalization": "G99", "Total Debt": "G101",
                 "Preferred Stock": "G102", "Enterprise Value": "G104",}
            
            yfinance_internal_map = {
                "Revenue (Sales)": ("revenue", "historical_projections"), "COGS (Cost of Goods Sold)": (None, "N/A"),
                "Gross Profit": ("Gross Profit", "income_stmt"), "SG&A (Selling, General & Administrative)": ("Selling General Administrative", "income_stmt"),
                "R&D (Research & Development)": ("Research Development", "income_stmt"), "Total Other Operating Components": (None, "N/A"),
                "EBITDA": ("EBITDA", "income_stmt"), "D&A (Depreciation & Amortization)": ("Depreciation", "cash_flow"),
                "Depreciation Expense": ("Depreciation And Amortization", "income_stmt"), "Amortization Expense": (None, "N/A"),
                "Operating Income (EBIT)": ("ebit", "historical_projections"), "Net Interest Expense (Income)": ("Interest Expense", "income_stmt"),
                "Interest Expense": ("Interest Expense", "income_stmt"), "Interest Income": ("Interest Income", "income_stmt"),
                "FX (Gain) Loss": (None, "N/A"), "Other Non-Operating (Income) Expenses": (None, "N/A"), 
                "Pre-Tax Income (EBT)": ("Pretax Income", "income_stmt"), "Tax Expense (Benefits)": ("Tax Provision", "income_stmt"),
                "Net Income": ("net_income", "historical_projections"), "EPS Basic": (None, "N/A"), "EPS Diluted": (None, "N/A"),
                "Basic Weighted Average Shares": ("Basic Average Shares", "income_stmt"), "Diluted Weighted Average Shares": ("Diluted Average Shares", "income_stmt"),
                "Cash & Cash Equivalents": ("Cash And Cash Equivalents", "balance_sheet"), "Short-Term Investments": ("Short Term Investments", "balance_sheet"),
                "Accounts Receivable": ("Accounts Receivable", "balance_sheet"), "Inventory": ("Inventory", "balance_sheet"),
                "Current Assets": ("Total Current Assets", "balance_sheet"), "Gross PP&E (Property, Plant and Equipment)": ("Gross PPE", "balance_sheet"),
                "Accumulated Depreciation": ("Accumulated Depreciation", "balance_sheet"), "Right-of-Use Assets": (None, "N/A"),
                "Intangibles": ("Intangible Assets", "balance_sheet"), "Goodwill": ("Goodwill", "balance_sheet"),
                "Non-Current Assets": ("Total Non Current Assets", "balance_sheet"), "Accounts Payable": ("Accounts Payable", "balance_sheet"),
                "Short-Term Borrowings": ("Short Term Debt", "balance_sheet"), "Current Portion of Lease Liabilities": (None, "N/A"),
                "Current Liabilities": ("Total Current Liabilities", "balance_sheet"), "Long-Term Borrowings": ("Long Term Debt", "balance_sheet"),
                "Long-Term Operating Lease Liabilities": (None, "N/A"), "Non-Current Liabilities": ("Total Non Current Liabilities", "balance_sheet"),
                "Non-Controlling Interest": ("Minority Interest", "balance_sheet"),
                "(Increase) Decrease in Accounts Receivable": ("Change In Receivables", "cash_flow"), "(Increase) Decrease in Inventories": ("Change In Inventory", "cash_flow"),
                "(Increase) Decrease in Pre-paid expeses and Other CA": ("Changes In Other Current Assets", "cash_flow"),
                "Increase (Decrease) in Accounts Payable": ("Change In Payables And Accrued Expense", "cash_flow"),
                "Increase (Decrease) in Accrued Revenues and Other CL": ("Changes In Other Current Liabilities", "cash_flow"),
                "Stock Based Compensation": ("Stock Based Compensation", "cash_flow"), "Operating Cash Flow": ("Operating Cash Flow", "cash_flow"),
                "Acquisition of Fixed & Intangibles": ("Capital Expenditure", "cash_flow"), "Disposal of Fixed & Intangibles": ("Sale Of PPE", "cash_flow"),
                "Acquisitions": ("Acquisitions And Divestitures", "cash_flow"), "Divestitures": (None, "N/A"),
                "Increase in LT Investment": ("Net Investment Purchase And Sale", "cash_flow"), "Decrease in LT Investment": (None, "N/A"),
                "Investing Cash Flow": ("Investing Cash Flow", "cash_flow"), "Debt Borrowing": ("Issuance Of Debt", "cash_flow"),
                "Debt Repayment": ("Repayment Of Debt", "cash_flow"), "Lease Payments": (None, "N/A"),
                "Dividends": ("Cash Dividends Paid", "cash_flow"), "Increase (Repurchase) of Shares": ("Repurchase Of Capital Stock", "cash_flow"),
                "Financing Cash Flow": ("Financing Cash Flow", "cash_flow"), "Effect of Foreign Exchange": ("Effect Of Exchange Rate Changes", "cash_flow"),
                "Market Capitalization": ("market_cap", "company_info"), "Total Debt": ("total_debt", "historical_only"),
                "Preferred Stock": ("Preferred Stock Equity", "balance_sheet"), "Enterprise Value": ("enterprise_value", "dcf_results"),}

            hist_cols_excel = ['G', 'H', 'I']; proj_cols_excel = ['J', 'K', 'L', 'M', 'N']
            num_hist_years_available = len(self.historical_years) 

            for excel_name, base_cell_str in field_cell_map.items():
                base_row_num = int(base_cell_str[1:])
                yfinance_key, source_type = yfinance_internal_map.get(excel_name, (None, "N/A"))

                # Default fill N/A for the entire row if mapping is N/A or key is missing for non-special types
                if source_type == "N/A" or (yfinance_key is None and source_type not in ["company_info", "dcf_results", "historical_only"]):
                    for col_idx in range(len(hist_cols_excel) + len(proj_cols_excel)):
                        target_col_letter = get_column_letter(column_index_from_string(hist_cols_excel[0]) + col_idx)
                        inputs_sheet[f"{target_col_letter}{base_row_num}"] = "N/A"
                    continue
                
                # Handle single current values (Market Cap, Enterprise Value)
                if excel_name in ["Market Capitalization", "Enterprise Value"]:
                    value_from_attr = None
                    if excel_name == "Market Capitalization": value_from_attr = self.company_info.get('market_cap')
                    elif excel_name == "Enterprise Value": value_from_attr = self.dcf_results.get('enterprise_value')
                    
                    target_cell_excel = f"{hist_cols_excel[-1]}{base_row_num}" # Col I
                    if isinstance(value_from_attr, (int, float)) and not pd.isna(value_from_attr):
                        inputs_sheet[target_cell_excel] = value_from_attr / 1000; inputs_sheet[target_cell_excel].number_format = '#,##0'
                    else: inputs_sheet[target_cell_excel] = "N/A"
                    # Fill other year cells for this specific item as N/A
                    for idx, col_ltr in enumerate(hist_cols_excel[:-1]): inputs_sheet[f"{col_ltr}{base_row_num}"] = "N/A"
                    for col_ltr in proj_cols_excel: inputs_sheet[f"{col_ltr}{base_row_num}"] = "N/A"
                    continue

                # Historical Data
                for i, col_letter in enumerate(hist_cols_excel):
                    cell_to_write = f"{col_letter}{base_row_num}"; raw_val = "N/A" 
                    if i < num_hist_years_available:
                        year_ts = self.historical_years[i] 
                        year_str_lookup = year_ts.strftime('%Y') 

                        df_to_use = None; current_raw_val = "N/A" 
                        if source_type == "historical_projections": current_raw_val = self.historical.get(year_str_lookup, {}).get(yfinance_key, "N/A")
                        elif source_type == "historical_only": current_raw_val = self.historical.get(year_str_lookup, {}).get(yfinance_key, "N/A")
                        elif source_type == "income_stmt": df_to_use = self.income_stmt
                        elif source_type == "balance_sheet": df_to_use = self.balance_sheet
                        elif source_type == "cash_flow": df_to_use = self.cash_flow
                        
                        if df_to_use is not None and not df_to_use.empty and yfinance_key in df_to_use.index and year_ts in df_to_use.columns:
                            val_from_df = df_to_use.loc[yfinance_key, year_ts]
                            if not pd.isna(val_from_df): current_raw_val = val_from_df

                        if excel_name == "Depreciation Expense" and (current_raw_val == "N/A" or pd.isna(current_raw_val) or current_raw_val == 0): 
                            if not self.cash_flow.empty and "Depreciation" in self.cash_flow.index and year_ts in self.cash_flow.columns:
                                raw_val_cf = self.cash_flow.loc["Depreciation", year_ts]
                                if not pd.isna(raw_val_cf) and raw_val_cf !=0: current_raw_val = raw_val_cf 
                        raw_val = current_raw_val 
                    
                    if isinstance(raw_val, (int, float)) and not pd.isna(raw_val):
                        inputs_sheet[cell_to_write] = raw_val / 1000; inputs_sheet[cell_to_write].number_format = '#,##0'
                    else: inputs_sheet[cell_to_write] = "N/A"

                # Projected Data
                if source_type == "historical_projections": 
                    for i, col_letter in enumerate(proj_cols_excel):
                        cell_to_write = f"{col_letter}{base_row_num}"; raw_val = "N/A"
                        if i < len(self.projection_years): 
                            proj_year_str = self.projection_years[i]
                            raw_val = self.projections.get(proj_year_str, {}).get(yfinance_key, "N/A")
                        if isinstance(raw_val, (int, float)) and not pd.isna(raw_val):
                            inputs_sheet[cell_to_write] = raw_val / 1000; inputs_sheet[cell_to_write].number_format = '#,##0'
                        else: inputs_sheet[cell_to_write] = "N/A"
                else: 
                    for col_letter in proj_cols_excel: inputs_sheet[f"{col_letter}{base_row_num}"] = "N/A"
            
            wb.save(output_path)
            print(f"Excel template populated: {output_path}")
            wb.close()
        except FileNotFoundError: print(f"Error: Template file '{template_path}' not found.")
        except KeyError as e: print(f"KeyError during Excel population: {e}. Check yfinance data for {self.ticker} or field name in mapping.")
        except Exception as e: print(f"An unexpected error occurred during Excel population: {str(e)}"); import traceback; traceback.print_exc()

def main():
    if len(sys.argv) < 2:
        print("Usage: python dcf_automator.py <TICKER> [template.xlsx] [output.xlsx]")
        return
    ticker = sys.argv[1]
    model = DCFModel(ticker)
    analysis_success = model.run_full_analysis()

    if len(sys.argv) == 4: 
        if analysis_success:
            template_path = sys.argv[2]
            output_path = sys.argv[3]
            if not os.path.exists(template_path):
                print(f"Error: Template file '{template_path}' does not exist.")
                return
            print(f"Analysis complete. Populating Excel template: {template_path} -> {output_path}")
            model.populate_excel_template(template_path, output_path)
        else: 
            print(f"DCF analysis failed for {ticker}. Excel population skipped.")
    elif analysis_success: 
         print(f"DCF analysis for {ticker} completed. No Excel paths provided, skipping Excel population.")
    # else: # Analysis failed and not 4 args
        # if len(sys.argv) != 2 : 
        #      print(f"DCF analysis failed for {ticker}. Also, incorrect arguments or paths provided for Excel population.")
        #      print("Usage for Excel: python dcf_automator.py <TICKER> <template.xlsx> <output.xlsx>")
        # # If only ticker was provided and analysis failed, run_full_analysis already printed errors.


if __name__ == "__main__":
    main()
