# You need to login to the Bloomberg Terminal for the script to work!
# Run it using the arrow on the top right.
# Enter the stock ticker while specifying the country and "Equity" in the end, 
# for example AAPL US or 000660 KS (the script will automatically add "Equity")

import blpapi
import openpyxl
import shutil
import os
import time
from datetime import datetime

def setup_bloomberg_session(ticker_symbol):
    options = blpapi.SessionOptions()
    options.setServerHost("localhost") 
    options.setServerPort(8194)        
    session = blpapi.Session(options)
    
    print(f"[INFO] Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("[WARNING] Failed to start Bloomberg session. Ensure Bloomberg Terminal is running and API is enabled.")
        return None
    if not session.openService("//blp/refdata"): 
        print("[WARNING] Failed to open Bloomberg reference data service.")
        session.stop()
        return None
    print("[INFO] Bloomberg session started successfully.")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=30):
    if not fields:
        print("[INFO] No fields to fetch in this batch.")
        return {}
        
    if len(fields) > 25: 
        print(f"[WARNING] Requesting {len(fields)} fields. Splitting into batches is handled by the calling function, but this single call exceeds typical safe limits if not batched.")

    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")
    
    security_for_request = f"{ticker} Equity" 
    request.getElement("securities").appendValue(security_for_request)
    
    parts = ticker.strip().split()
    country_code = ""
    if len(parts) > 1 and len(parts[-1]) == 2 and parts[-1].isalpha():
        country_code = parts[-1].upper()

    if country_code and country_code != "US":
        print(f"[INFO] Applying currency override to USD for non-US ticker: {ticker} (Country: {country_code})")
        request.set("currency", "USD")
    elif country_code == "US":
        print(f"[INFO] Ticker {ticker} is US-based. No currency override applied (data typically in USD).")
    else:
        print(f"[INFO] No specific non-US country code identified or format unrecognized for {ticker}. No currency override applied. Data will be in local currency.")

    for field in fields:
        request.getElement("fields").appendValue(field)
    
    request.set("periodicitySelection", "YEARLY") 
    request.set("startDate", f"{start_year}0101") 
    request.set("endDate", f"{end_year}1231")   
    
    print(f"[DEBUG] Sending request for {security_for_request} with fields: {fields}")
    session.sendRequest(request)
    
    data = {field: {} for field in fields} 
    invalid_fields = [] 
    start_time = time.time()
    
    while True: 
        if time.time() - start_time >= timeout:
            print(f"[WARNING] Timed out waiting for complete response for {security_for_request} after {timeout}s.")
            for field_id_timeout in fields:
                if field_id_timeout not in invalid_fields:
                    for year_timeout_val in range(start_year, end_year + 1):
                        if year_timeout_val not in data.get(field_id_timeout, {}):
                            data.setdefault(field_id_timeout, {})[year_timeout_val] = "N/A (Timeout)"
            break 

        event = session.nextEvent(500) 
        
        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"[DEBUG] Bloomberg nextEvent() timeout for {security_for_request}. Continuing to wait for data...")
            continue 

        for msg in event:
            
            if msg.hasElement("responseError"):
                error = msg.getElement("responseError")
                error_message = error.getElement("message").getValue()
                print(f"[ERROR] Bloomberg API responseError for {security_for_request}: {error_message}")
                for f_id_err in fields:
                    if f_id_err not in invalid_fields: invalid_fields.append(f_id_err)
                continue 

            if not msg.hasElement("securityData"):
                print(f"[WARNING] No securityData element in message for {security_for_request}.")
                continue 

            security_data = msg.getElement("securityData") 
            
            if security_data.hasElement("securityError"):
                sec_error = security_data.getElement("securityError")
                error_msg_sec = sec_error.getElement("message").getValueAsString()
                print(f"[ERROR] Security error for '{security_data.getElementValue('security')}': {error_msg_sec}")
                for f_id_sec_err in fields:
                    if f_id_sec_err not in invalid_fields: invalid_fields.append(f_id_sec_err)
                continue 

            if security_data.hasElement("fieldExceptions"):
                field_exceptions = security_data.getElement("fieldExceptions")
                for j in range(field_exceptions.numValues()):
                    field_error = field_exceptions.getValue(j)
                    invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                    error_info = field_error.getElement("errorInfo")
                    error_message_field = error_info.getElement("message").getValueAsString()
                    field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field (Not in map)")
                    print(f"[WARNING] Invalid Bloomberg field: '{invalid_field_id}' (Excel Name: '{field_name_display}') for {security_for_request}. Error: {error_message_field}")
                    if invalid_field_id not in invalid_fields:
                        invalid_fields.append(invalid_field_id)
            
            if not security_data.hasElement("fieldData"):
                print(f"[WARNING] No fieldData element in securityData for {security_for_request}. This might occur if all fields were invalid or no data for the period.")
                for f_id_no_data in fields:
                    if f_id_no_data not in invalid_fields:
                        for year_val_nd in range(start_year, end_year + 1):
                             data.setdefault(f_id_no_data, {})[year_val_nd] = "N/A (No fieldData)"
                continue 

            field_data_array = security_data.getElement("fieldData") 
            # print(f"[DEBUG] Number of fieldData entries (dates): {field_data_array.numValues()}")
            
            for k in range(field_data_array.numValues()): 
                datum = field_data_array.getValue(k)
                if not datum.hasElement("date"):
                    print("[WARNING] fieldData entry missing 'date' element. Skipping this entry.")
                    continue
                date_obj = datum.getElement("date").getValueAsDatetime()
                year = date_obj.year
                
                for field_id in fields:
                    if field_id in invalid_fields:
                        data.setdefault(field_id, {})[year] = "N/A (Invalid Field)"
                        continue
                    
                    if datum.hasElement(field_id):
                        try:
                            value = datum.getElement(field_id).getValueAsFloat()
                            data.setdefault(field_id, {})[year] = value
                        except blpapi.exception.ElementErrorException:
                            try:
                                value_str = datum.getElement(field_id).getValueAsString()
                                data.setdefault(field_id, {})[year] = value_str
                                print(f"[DEBUG] Field {field_id} for year {year} for {security_for_request} is not a float, stored as string: {value_str}")
                            except Exception as e_str:
                                print(f"[WARNING] Could not get value for field {field_id} for year {year} for {security_for_request} (even as string): {e_str}")
                                data.setdefault(field_id, {})[year] = "N/A (Error extracting)"
                        except Exception as e_gen: 
                             print(f"[WARNING] Error converting field {field_id} for year {year} for {security_for_request}: {e_gen}")
                             data.setdefault(field_id, {})[year] = "N/A (Conversion Error)"
                    else:
                        if year not in data.get(field_id, {}):
                            data.setdefault(field_id, {})[year] = None 
            
        if event.eventType() == blpapi.Event.RESPONSE: 
            print(f"[INFO] Received final response for batch for {security_for_request}.")
            for field_id_fill in fields:
                if field_id_fill not in invalid_fields:
                    for year_fill_val in range(start_year, end_year + 1):
                        if year_fill_val not in data.get(field_id_fill, {}): 
                            data.setdefault(field_id_fill, {})[year_fill_val] = "N/A (Missing)"
            break 

        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("[ERROR] Bloomberg session terminated unexpectedly. Further requests may fail.")
                    return None 
                print(f"[INFO] Session/Service Status update: {msg.toString()}")
        else:
            print(f"[INFO] Received unhandled Bloomberg event type: {event.eventType()}")

    if not any(data.get(field) for field in data): 
        print(f"[WARNING] No data successfully retrieved for any requested field for {ticker} in this batch.")
    
    if invalid_fields:
        print(f"[INFO] Bloomberg fields skipped or marked N/A due to invalidity/errors for {security_for_request}: {invalid_fields}")
    
    return data

def calculate_derived_metrics(data, start_year=2014, end_year=2024):
    derived = {
        "DSO": {} 
    }
    
    def get_val(source_field_code, year, default=0.0):
        val = data.get(source_field_code, {}).get(year)
        if isinstance(val, (int, float)):
            return val
        return default

    for year in range(start_year, end_year + 1):
        revenue = get_val("SALES_REV_TURN", year) 
        ar = get_val("BS_ACCT_NOTE_RCV", year)   
        
        if revenue is not None and ar is not None and not isinstance(revenue, str) and not isinstance(ar, str):
            derived["DSO"][year] = (ar / revenue * 365) if revenue != 0 else 0.0
        else:
            derived["DSO"][year] = "N/A (Missing inputs)"
            
    return derived

field_map = {
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "ARDR_DEPRECIATION_AMORTIZATION", "statement": "IS"},
    "Depreciation Expense": {"source": "BDH", "field": "ARDR_DEPRECIATION_EXP", "statement": "IS"},
    "Amortization Expense": {"source": "BDH", "field": "ARDR_AMORT_EXP", "statement": "IS"},
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPER_INC", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS", "statement": "IS"},
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_AVG_NUM_SH_FOR_EPS", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_SH_FOR_DILUTED_EPS", "statement": "IS"},
    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS"}, 
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS"},
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS"},
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLES", "statement": "BS"},
    "Goodwill": {"source": "BDH", "field": "BS_GOODWILL", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSET", "statement": "BS"},
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "SHORT_TERM_DEBT_DETAILED", "statement": "BS"},
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "ST_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"}, 
    "Long-Term Borrowings": {"source": "BDH", "field": "LONG_TERM_BORROWINGS_DETAILED", "statement": "BS"},
    "Long-Term Operating Lease Liabilities": {"source": "BDH", "field": "LT_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Non-Current Liabilities": {"source": "BDH", "field": "NON_CUR_LIAB", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"}, 
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHANGE_IN_INVENTORIES", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Pre-paid expeses and Other CA": {"source": "BDH", "field": "CF_CHANGE_IN_OTH_CUR_ASSETS", "statement": "CF", "section": "Operating"}, 
    "Increase (Decrease) in Accounts Payable": {"source": "BDH", "field": "CF_CHANGE_IN_ACCOUNTS_PAYABLE", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accrued Revenues and Other CL": {"source": "BDH", "field": "CF_CHANGE_IN_OTH_CUR_LIAB", "statement": "CF", "section": "Operating"}, 
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMPENSATION", "statement": "CF", "section": "Operating"},
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "ACQUIS_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"}, 
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "DISPOSAL_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_CASH_FOR_DIVESTITURES", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_INCR_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECR_INVEST", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PROCEEDS", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PAYMENT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "PROC_FR_REPURCH_EQTY_DETAILED", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CFF_ACTIVITIES_DETAILED", "statement": "CF", "section": "Financing"},
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FOREIGN_EXCHANGES", "statement": "CF", "section": "All"},
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"}, 
    "Total Debt": {"source": "BDH", "field": "SHORT_AND_LONG_TERM_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "PFD_EQTY_HYBRID_CAPITAL", "statement": "BS"},
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"}, 
    "Total Borrowings": {"source": "BDH", "field": "TOT_BORROWINGS", "statement": "BS"},
    "Total Leases": {"source": "BDH", "field": "TOT_LEASE_LIAB", "statement": "BS"},
    "Net Debt": {"source": "BDH", "field": "NET_DEBT", "statement": "BS"},
    "Effective Tax Rate": {"source": "BDH", "field": "EFF_TAX_RATE", "statement": "IS"}, 
    "DSO": {"source": "derived", "field": "DSO", "statement": "IS"}, 
}
#add per request 
#right of use assets

field_cell_map = {
    "Revenue (Sales)": "G6", "COGS (Cost of Goods Sold)": "G7", "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9", "R&D (Research & Development)": "G10",
    "Other Operating (Income) Expenses": "G11", "EBITDA": "G12", "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14", "Amortization Expense": "G15", "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17", "Interest Expense": "G18", "Interest Income": "G19",
    "FX (Gain) Loss": "G20", "Other Non-Operating (Income) Expenses": "G21", "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23", "Net Income": "G24", "EPS Basic": "G25", "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27", "Diluted Weighted Average Shares": "G28",
    "Effective Tax Rate": "G29", 
    "Cash & Cash Equivalents": "G33", "Short-Term Investments": "G34", "Accounts Receivable": "G35",
    "Inventory": "G36", "Current Assets": "G38", "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41", "Intangibles": "G43", "Goodwill": "G44",
    "Non-Current Assets": "G47", "Accounts Payable": "G49", "Short-Term Borrowings": "G51",
    "Current Portion of Lease Liabilities": "G52", "Current Liabilities": "G54",
    "Long-Term Borrowings": "G56", "Long-Term Operating Lease Liabilities": "G57",
    "Non-Current Liabilities": "G59", "Non-Controlling Interest": "G62", 
    "Total Debt": "G63", "Preferred Stock": "G64", "Total Borrowings": "G65", "Total Leases": "G66", "Net Debt": "G67",
    "D&A (Depreciation & Amortization)": "G70", 
    "(Increase) Decrease in Accounts Receivable": "G71", "(Increase) Decrease in Inventories": "G72",
    "(Increase) Decrease in Pre-paid expeses and Other CA": "G73",
    "Increase (Decrease) in Accounts Payable": "G74",
    "Increase (Decrease) in Accrued Revenues and Other CL": "G75", 
    "Stock Based Compensation": "G76", "Operating Cash Flow": "G78", 
    "Acquisition of Fixed & Intangibles": "G80", "Disposal of Fixed & Intangibles": "G81", 
    "Acquisitions": "G83", "Divestitures": "G84", "Increase in LT Investment": "G85",
    "Decrease in LT Investment": "G86", "Investing Cash Flow": "G88", 
    "Debt Borrowing": "G90", "Debt Repayment": "G91", "Dividends": "G92", 
    "Increase (Repurchase) of Shares": "G93", "Financing Cash Flow": "G95", 
    "Effect of Foreign Exchange": "G96",
    "Market Capitalization": "G101", "Enterprise Value": "G102",
    "DSO": "G105", 
}

def filter_field_map_for_task(task_name, current_field_map):
    statement_code = task_name
    
    allowed_statements = ["IS", "BS", "CF"]
    if statement_code not in allowed_statements:
        raise ValueError(f"Invalid statement code '{statement_code}'. Must be one of {allowed_statements}.")
    
    task_specific_configs = {} 
    for name, config in current_field_map.items():
        if config["statement"] == statement_code:
            task_specific_configs[name] = config
    
    required_bdh_for_derived_metrics = set()
    for name, config in task_specific_configs.items(): 
        if config["source"] == "derived":
            if config["field"] == "DSO": 
                required_bdh_for_derived_metrics.add("BS_ACCT_NOTE_RCV") 
                required_bdh_for_derived_metrics.add("SALES_REV_TURN")   

    for bdh_field_code_needed in required_bdh_for_derived_metrics:
        found_in_task = False
        for _, existing_config in task_specific_configs.items():
            if existing_config.get("field") == bdh_field_code_needed and existing_config.get("source") == "BDH":
                found_in_task = True
                break
        if not found_in_task:
            for global_name, global_config in current_field_map.items():
                if global_config.get("field") == bdh_field_code_needed and global_config.get("source") == "BDH":
                    pass 
    return task_specific_configs

def batch_fields(fields_to_fetch, batch_size=25):
    unique_fields = sorted(list(set(fields_to_fetch))) 
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index):
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    try:
        col_str = "".join(filter(str.isalpha, base_cell_ref))
        row_num = int("".join(filter(str.isdigit, base_cell_ref)))
        start_col_idx = openpyxl.utils.column_index_from_string(col_str) 
        
        target_cells = []
        for i in range(num_years):
            target_col_letter = get_column_letter_from_index(start_col_idx + i)
            target_cells.append(f"{target_col_letter}{row_num}")
        return target_cells
    except ValueError as e:
        print(f"[ERROR] Invalid base cell reference '{base_cell_ref}': {e}")
        raise 

def populate_valuation_model(template_path, output_path, ticker_symbol, current_field_map, current_field_cell_map):
    if not os.path.exists(template_path):
        print(f"[ERROR] Template file '{template_path}' not found. Please ensure it's in the script's directory or provide the full path.")
        raise FileNotFoundError(f"Template file {template_path} not found.")

    try:
        shutil.copy(template_path, output_path)
        print(f"[INFO] Copied template '{template_path}' to output file '{output_path}'.")
    except Exception as e_copy:
        print(f"[ERROR] Failed to copy template to output path: {e_copy}")
        raise

    try:
        wb = openpyxl.load_workbook(output_path) 
    except Exception as e_load:
        print(f"[ERROR] Failed to load the copied workbook from '{output_path}': {e_load}")
        raise

    if "Inputs" not in wb.sheetnames:
        print("[ERROR] 'Inputs' sheet not found in the workbook. Please check the template.")
        raise ValueError("'Inputs' sheet not found in the template file.")
    ws = wb["Inputs"]
    
    data_years = list(range(2014, 2024 + 1)) 
    num_data_years = len(data_years)
    
    all_fetched_bdh_data = {} 
    
    global_bberg_code_to_excel_name_map = {
        config["field"]: name 
        for name, config in current_field_map.items() 
        if config.get("source") == "BDH" and "field" in config
    }
    
    all_bdh_fields_to_fetch_codes = set()
    for excel_name, config in current_field_map.items():
        if config.get("source") == "BDH" and "field" in config:
            all_bdh_fields_to_fetch_codes.add(config["field"])
        elif config.get("source") == "derived":
            if config["field"] == "DSO":
                all_bdh_fields_to_fetch_codes.add("BS_ACCT_NOTE_RCV")
                all_bdh_fields_to_fetch_codes.add("SALES_REV_TURN")

    if not all_bdh_fields_to_fetch_codes:
        print("[WARNING] No Bloomberg (BDH) fields identified for fetching across all configurations. Check field_map.")
        wb.save(output_path)
        return

    print(f"\n[PHASE] Starting data fetching for ticker: {ticker_symbol}")
    print(f"[INFO] Total unique BDH fields to fetch: {len(all_bdh_fields_to_fetch_codes)}")
    
    field_batches = batch_fields(list(all_bdh_fields_to_fetch_codes), batch_size=25)
    print(f"[INFO] Split into {len(field_batches)} batches for fetching.")
    
    session = None 
    try:
        session = setup_bloomberg_session(ticker_symbol)
        if not session:
            print(f"[ERROR] Failed to start Bloomberg session for {ticker_symbol}. Cannot fetch data.")
            wb.save(output_path) 
            raise ConnectionError("Failed to establish Bloomberg session.")

        for batch_idx, current_batch_bberg_codes in enumerate(field_batches):
            print(f"    [BATCH] Processing batch {batch_idx + 1}/{len(field_batches)} with {len(current_batch_bberg_codes)} fields.")
            
            batch_data_fetched = fetch_bloomberg_data(
                session, 
                ticker_symbol, 
                current_batch_bberg_codes, 
                global_bberg_code_to_excel_name_map, 
                start_year=data_years[0], 
                end_year=data_years[-1]
            )
            
            if batch_data_fetched is None: 
                print(f"    [ERROR] Critical error fetching data for batch {batch_idx + 1}. Aborting further fetches.")
                raise ConnectionAbortedError("Bloomberg session terminated or critical fetch error.")
            
            elif batch_data_fetched: 
                for field_code, yearly_data in batch_data_fetched.items():
                    if field_code not in all_fetched_bdh_data:
                        all_fetched_bdh_data[field_code] = {}
                    for year, value in yearly_data.items():
                        if value is not None: 
                             all_fetched_bdh_data[field_code][year] = value
                        elif year not in all_fetched_bdh_data[field_code]: 
                             all_fetched_bdh_data[field_code][year] = value 
                print(f"    [SUCCESS] Fetched data for batch {batch_idx + 1}. {len(batch_data_fetched)} fields processed from batch.")
            else:
                print(f"    [INFO] No data returned for batch {batch_idx + 1}, or batch was empty/all fields invalid.")
    
    except Exception as e_fetch:
        print(f"[ERROR] An unexpected error occurred during the data fetching phase: {e_fetch}")
    finally:
        if session:
            try:
                session.stop()
                print("[INFO] Bloomberg session stopped.")
            except Exception as e_stop:
                print(f"[WARNING] Error stopping Bloomberg session: {e_stop}")
    
    print(f"\n[PHASE] Completed all data fetching attempts.")

    print(f"\n[PHASE] Calculating derived metrics...")
    all_derived_data = calculate_derived_metrics(all_fetched_bdh_data, start_year=data_years[0], end_year=data_years[-1])
    print("[INFO] Derived metrics calculated.")
    
    print(f"\n[PHASE] Writing all data to Excel sheet '{ws.title}'...")
    
    for excel_name, config in current_field_map.items():
        if excel_name.startswith("__dep_"): 
            continue
            
        base_cell_ref = current_field_cell_map.get(excel_name)
        if not base_cell_ref:
            print(f"[WARNING] No Excel cell mapping found for '{excel_name}'. Skipping this item.")
            continue
            
        try:
            target_cells_for_item = get_target_cells_for_years(base_cell_ref, num_data_years)
        except Exception as e_cell_calc:
            print(f"[ERROR] Could not calculate target cells for '{excel_name}' with base '{base_cell_ref}': {e_cell_calc}. Skipping.")
            continue
        
        data_source_for_item = {}
        if config["source"] == "BDH":
            bberg_field_code = config.get("field")
            if not bberg_field_code:
                print(f"[WARNING] BDH item '{excel_name}' has no Bloomberg field code defined. Skipping.")
                continue
            data_source_for_item = all_fetched_bdh_data.get(bberg_field_code, {})

        elif config["source"] == "derived":
            derived_field_key = config.get("field") 
            if not derived_field_key:
                print(f"[WARNING] Derived item '{excel_name}' has no derived field key defined. Skipping.")
                continue
            data_source_for_item = all_derived_data.get(derived_field_key, {})
        else:
            print(f"[INFO] Item '{excel_name}' has unknown source '{config['source']}'. Skipping.")
            continue

        for i, year in enumerate(data_years):
            cell_ref = target_cells_for_item[i]
            raw_value = data_source_for_item.get(year) 
            
            display_value = raw_value
            
            if raw_value is None:
                display_value = "N/A (Missing)" 
            
            if isinstance(raw_value, (int, float)):
                ws[cell_ref] = raw_value
                ws[cell_ref].number_format = "#,##0.000" 
                if "EPS" in excel_name or "DSO" in excel_name or "Rate" in excel_name: 
                     ws[cell_ref].number_format = "0.00" 
            elif isinstance(raw_value, str): 
                ws[cell_ref] = raw_value
            else: 
                ws[cell_ref] = "N/A" if raw_value is None else str(raw_value) 
    
    try:
        wb.save(output_path)
        print(f"\n[SUCCESS] Valuation model populated and saved to '{output_path}'")
    except Exception as e_save:
        print(f"[ERROR] Failed to save the final workbook to '{output_path}': {e_save}")
        print(f"[INFO] The file might be open or permissions might be an issue.")


if __name__ == "__main__":
    print("-" * 70)
    print("Bloomberg Data to Excel Valuation Model Populator")
    print("-" * 70)

    excel_template_path = "LIS_Valuation_Empty.xlsx" 

    try:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_folder_name = "Bloomberg_Valuation_Reports" 
        output_directory = os.path.join(desktop_path, output_folder_name)

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"[SETUP] Created output directory: '{output_directory}'")
        else:
            print(f"[SETUP] Output directory already exists: '{output_directory}'")
    except Exception as e_path:
        print(f"[WARNING] Could not create/access Desktop output directory: {e_path}")
        print("[WARNING] Defaulting to saving in the script's current directory.")
        output_directory = "." 
    
    ticker_input = ""
    while not ticker_input:
        raw_input_str = input("Enter Ticker Symbol (e.g., AAPL US or 000660 KS): ").strip()
        if raw_input_str and any(char.isalnum() for char in raw_input_str):
            ticker_input = raw_input_str.upper() 
        else:
            print("[VALIDATION] Ticker symbol cannot be empty and must contain alphanumeric characters. Please try again.")
    
    safe_ticker_filename = ticker_input.replace(" ", "_").replace("/", "_")
    output_file_name = f"{safe_ticker_filename}_Valuation_Model_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    final_output_path = os.path.join(output_directory, output_file_name)
    
    print(f"\n[SETUP] Template: '{excel_template_path}'")
    print(f"[SETUP] Final Output will be: '{final_output_path}'")
    print(f"[SETUP] Ticker for Bloomberg: '{ticker_input}' (Script will append ' Equity')")
    
    try:
        print("\nStarting the data population process...\n")
        populate_valuation_model(
            template_path=excel_template_path,
            output_path=final_output_path,
            ticker_symbol=ticker_input, 
            current_field_map=field_map, 
            current_field_cell_map=field_cell_map 
        )
        print("\nProcess completed.")
    except FileNotFoundError as e_fnf:
        print(f"[CRITICAL ERROR] File not found: {e_fnf}")
        print("[INFO] Please ensure the Excel template exists and paths are correct.")
    except ConnectionError as e_conn:
        print(f"[CRITICAL ERROR] Bloomberg connection issue: {e_conn}")
        print("[INFO] Ensure Bloomberg Terminal is running and API is properly configured.")
    except ConnectionAbortedError as e_aborted:
        print(f"[CRITICAL ERROR] Bloomberg connection aborted during fetch: {e_aborted}")
    except blpapi.exception.BlpException as e_blp:
        print(f"[CRITICAL BLPAPI ERROR] A Bloomberg API exception occurred: {e_blp}")
    except Exception as e_main:
        print(f"[CRITICAL UNEXPECTED ERROR] An unexpected error occurred in the main process: {e_main}")
        import traceback
        print("\n--- Traceback ---")
        traceback.print_exc()
        print("--- End Traceback ---\n")
    finally:
        print("\nScript execution finished.")
