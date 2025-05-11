import blpapi
import openpyxl
import shutil
import os
import numpy as np
import time
from datetime import datetime

def setup_bloomberg_session(ticker_symbol):
    """Initialize Bloomberg API session with detailed logging."""
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)
    
    print(f"[INFO] Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("[WARNING] Failed to start Bloomberg session. Ensure Bloomberg Terminal is running.")
        return None
    if not session.openService("//blp/refdata"):
        print("[WARNING] Failed to open Bloomberg reference data service.")
        session.stop() # Stop session if service opening fails
        return None
    print("[INFO] Bloomberg session started successfully.")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=30): # Increased timeout
    """Fetch historical data from Bloomberg with timeout and error handling."""
    if not fields: # Skip if no fields in batch
        print("[INFO] No fields in the current batch to fetch.")
        return {}
        
    if len(fields) > 25:
        # This check is more of a safeguard; batching should prevent this.
        print(f"[ERROR] Too many fields ({len(fields)}). Bloomberg API limit is 25 fields per request. This should be handled by batching.")
        raise ValueError(f"Too many fields ({len(fields)}). Bloomberg API limit is 25 fields per request.")
    
    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")
    security = f"{ticker} US Equity"
    request.getElement("securities").appendValue(security)
    for field in fields:
        request.getElement("fields").appendValue(field)
    request.set("periodicitySelection", "YEARLY")
    request.set("startDate", f"{start_year}0101")
    request.set("endDate", f"{end_year}1231")
    
    print(f"[DEBUG] Sending request for {security} with fields: {fields}")
    session.sendRequest(request)
    
    data = {field: {} for field in fields}
    invalid_fields = []
    start_time = time.time()
    
    while time.time() - start_time < timeout:
        event = session.nextEvent(500) # 500ms timeout for nextEvent
        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"[DEBUG] Bloomberg event timeout for {security}. Continuing to wait for response...")
            continue # Continue waiting until overall timeout

        if event.eventType() in [blpapi.Event.RESPONSE, blpapi.Event.PARTIAL_RESPONSE]:
            for msg in event:
                # print(f"[DEBUG] Received message: {msg}") # Can be very verbose
                if msg.hasElement("responseError"):
                    error = msg.getElement("responseError")
                    error_message = error.getElement("message").getValue()
                    print(f"[ERROR] Bloomberg API error for {security}: {error_message}")
                    # Depending on the error, you might want to raise it or handle it
                    # For now, we'll let it proceed to see if partial data is available or if it's a final response
                    # raise ValueError(f"Bloomberg API error: {error_message}") # Option to fail fast

                if not msg.hasElement("securityData"):
                    print(f"[WARNING] No securityData element in response for {security}. Check ticker or data availability.")
                    continue # Process next message if any
                
                security_data_array = msg.getElement("securityData")
                # Bloomberg can return multiple securityData elements if multiple securities were requested,
                # but here we request one at a time.
                if security_data_array.numValues() == 0:
                    print(f"[WARNING] securityData array is empty for {security}.")
                    continue

                security_data = security_data_array.getValueAsElement(0) # Assuming one security per request
                
                # Check for field exceptions (invalid fields)
                if security_data.hasElement("fieldExceptions"):
                    field_exceptions = security_data.getElement("fieldExceptions")
                    for i in range(field_exceptions.numValues()):
                        field_error = field_exceptions.getValueAsElement(i)
                        invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                        error_info = field_error.getElement("errorInfo").getElement("message").getValueAsString()
                        field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field")
                        print(f"[WARNING] Invalid Bloomberg field: '{invalid_field_id}' (mapped to '{field_name_display}') for {security}. Error: {error_info}")
                        if invalid_field_id not in invalid_fields:
                            invalid_fields.append(invalid_field_id)
                
                if not security_data.hasElement("fieldData"):
                    print(f"[WARNING] No fieldData element in securityData for {security}.")
                    continue

                field_data_array = security_data.getElement("fieldData")
                for i in range(field_data_array.numValues()): # Iterate through historical data points (dates)
                    datum = field_data_array.getValueAsElement(i)
                    date_obj = datum.getElement("date").getValueAsDatetime()
                    year = date_obj.year
                    for field_id in fields:
                        if field_id in invalid_fields:
                            data[field_id][year] = "N/A (Invalid Field)" # Mark as N/A
                            continue
                        if datum.hasElement(field_id):
                            try:
                                value = datum.getElement(field_id).getValueAsFloat() # Assuming numeric data
                                data[field_id][year] = value
                            except blpapi.exception.ElementErrorException:
                                # Handle cases where field might not be float (e.g. #N/A History)
                                try:
                                   value_str = datum.getElement(field_id).getValueAsString()
                                   data[field_id][year] = value_str # Store as string if not float
                                   print(f"[DEBUG] Field {field_id} for year {year} for {security} is not a float, stored as string: {value_str}")
                                except Exception as e_str:
                                   print(f"[WARNING] Could not get value for field {field_id} for year {year} for {security}: {e_str}")
                                   data[field_id][year] = "N/A (Error)"
                        else:
                            # Field present in request but not in this specific dated entry
                            # This can happen if data for that field/year combination is missing
                            if year not in data[field_id]: # Initialize if not already set by another message
                                data[field_id][year] = None # Or some other placeholder like "N/A"
                                # print(f"[DEBUG] Field {field_id} not present in datum for year {year} for {security}.")


        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("[ERROR] Bloomberg session terminated unexpectedly.")
                    # Consider how to handle this: retry, raise error, etc.
                    # For now, returning None to indicate failure for this batch.
                    return None 
                # You can add more detailed status message handling here
                # print(f"[DEBUG] Session/Service Status: {msg}")

        if event.eventType() == blpapi.Event.RESPONSE: # Final response for the request
            print(f"[INFO] Received final response for batch for {security}.")
            break # Exit while loop
    
    if time.time() - start_time >= timeout and event.eventType() != blpapi.Event.RESPONSE:
        print(f"[WARNING] Timed out waiting for complete response for {security} after {timeout}s.")

    if not any(data[field] for field in data if data[field]): # Check if any data was actually populated
        print(f"[WARNING] No data received for any requested field for {ticker} in this batch.")
    
    if invalid_fields:
        print(f"[INFO] Bloomberg fields skipped or marked N/A due to invalidity for {security}: {invalid_fields}")
    
    return data


def calculate_derived_metrics(data, start_year=2014, end_year=2024):
    """Calculate derived metrics like DSO, DIH, DPO, and Net Working Capital changes."""
    derived = {
        "Changes in Net Working Capital": {},
        "DSO": {},
        "DIH": {},
        "DPO": {},
        "Net Cash from Investments & Acquisitions": {},
        "Increase (Decrease) in Other": {}
    }
    
    # Helper to safely get data, returning 0 if key or year is missing or value is not numeric
    def get_val(source_field, year, default=0):
        val = data.get(source_field, {}).get(year)
        if isinstance(val, (int, float)):
            return val
        # print(f"[DEBUG] Derived metric calc: Value for {source_field}/{year} is missing or not numeric ({val}). Using default {default}.")
        return default

    for year in range(start_year, end_year + 1):
        # Changes in Net Working Capital
        # NWC = Current Assets - Current Liabilities
        # Change in NWC = NWC_current_year - NWC_previous_year
        if year > start_year: # Requires previous year data
            nwc_t = get_val("BS_CUR_ASSET_REPORT", year) - get_val("BS_CUR_LIAB", year)
            nwc_t1 = get_val("BS_CUR_ASSET_REPORT", year - 1) - get_val("BS_CUR_LIAB", year - 1)
            # Only record if both current and previous year assets/liabilities were found (non-zero)
            if (get_val("BS_CUR_ASSET_REPORT", year) != 0 or get_val("BS_CUR_LIAB", year) != 0) and \
               (get_val("BS_CUR_ASSET_REPORT", year - 1) != 0 or get_val("BS_CUR_LIAB", year - 1) != 0):
                 derived["Changes in Net Working Capital"][year] = nwc_t - nwc_t1

        # DSO, DIH, DPO
        revenue = get_val("SALES_REV_TURN", year)
        cogs = get_val("IS_COG_AND_SERVICES_SOLD", year)
        ar = get_val("BS_ACCT_NOTE_RCV", year)
        inv = get_val("BS_INVENTORIES", year)
        ap = get_val("ACCT_PAYABLE_ACCRUALS_DETAILED", year)

        derived["DSO"][year] = (ar / revenue * 365) if revenue else 0
        derived["DIH"][year] = (inv / cogs * 365) if cogs else 0
        derived["DPO"][year] = (ap / cogs * 365) if cogs else 0
        
        # Net Cash from Investments & Acquisitions
        # Sum of cash for acquisitions, divestitures, and other investing activities
        cf_acq = get_val("CF_CASH_FOR_ACQUIS_SUBSIDIARIES", year)
        cf_div = get_val("CF_CASH_FOR_DIVESTURES", year)
        cf_oth_inv = get_val("OTHER_INVESTING_ACT_DETAILED", year)
        derived["Net Cash from Investments & Acquisitions"][year] = cf_acq + cf_div + cf_oth_inv
        
        # Increase (Decrease) in Other (Operating CF section)
        # This is a balancing item. Original formula:
        # Change in NWC - (Change in AR + Change in Inventory - Change in AP)
        # Simplified: Change in NWC - Change in AR - Change in Inv + Change in AP (if AP is positive for increase)
        # Bloomberg fields for CF are already changes.
        # So, Increase (Decrease) in Other = Change in NWC - ( (Increase)/Decrease AR + (Increase)/Decrease Inv + Increase/(Decrease) AP from CF)
        # Note: CF_ACCT_RCV_UNBILLED_REV is -(Increase in AR). So, -CF_ACCT_RCV_UNBILLED_REV is (Increase in AR)
        
        chg_nwc_val = derived.get("Changes in Net Working Capital", {}).get(year) # Use already calculated Chg NWC
        
        # CF fields are typically reported as:
        # CF_ACCT_RCV_UNBILLED_REV: Negative if AR increases (cash outflow)
        # CF_CHAGE_IN_INVENTORIES: Negative if Inventory increases (cash outflow)
        # ACCT_PAYABLE_ACCRUALS_DETAILED (from BS) is a balance. We need change in AP from CF.
        # Let's assume there's a CF field for change in AP, or it's part of "Other Operating Adjustments"
        # The provided field map has "ACCT_PAYABLE_ACCRUALS_DETAILED" for BS, not directly for CF change.
        # The original formula used BS AP. Let's stick to that for now, implying it's a component of NWC.
        # The formula seems to be: Chg NWC - (CF_delta_AR + CF_delta_Inv + CF_delta_AP)
        # where CF_delta_AR is the cash effect of AR change.
        # If using direct CF fields:
        # CF_ACCT_RCV_UNBILLED_REV (is -(Increase in AR))
        # CF_CHAGE_IN_INVENTORIES (is -(Increase in Inventory))
        # Need a CF_CHANGE_IN_PAYABLES. If not available, this metric might be hard to reconcile directly.

        # The user's original formula was:
        # derived["Changes in Net Working Capital"][year] -
        # (data["CF_ACCT_RCV_UNBILLED_REV"][year] +
        #  data["CF_CHAGE_IN_INVENTORIES"][year] +
        #  data["ACCT_PAYABLE_ACCRUALS_DETAILED"][year]) <--- This was BS AP, not change in AP.

        # Let's assume "Increase (Decrease) in Other" is a plug to reconcile Operating CF.
        # Op CF = Net Income + D&A + Stock Comp + Change NWC (simplified)
        # Or Net Income + D&A + Adjustments for non-cash items + Changes in Op Assets/Liabilities
        # The fields CF_ACCT_RCV_UNBILLED_REV, CF_CHAGE_IN_INVENTORIES are part of "Changes in Op Assets/Liabilities"
        # If "Increase (Decrease) in Other" is meant to be the remaining part of "Changes in Op Assets/Liabilities"
        # after AR, Inv, AP are accounted for from NWC, then:
        # Increase (Decrease) in Other = Change in NWC - (Change in AR + Change in Inv - Change in AP)
        # where Change in AR (asset) is negative for CF if it increases.
        # This derived metric seems complex without a clear standard definition or all specific CF change fields.
        # For now, let's use the user's prior logic if the fields are available, but acknowledge it might need review.
        
        if chg_nwc_val is not None: # Ensure Change in NWC was calculated
            cf_ar_change = get_val("CF_ACCT_RCV_UNBILLED_REV", year) # Cash effect from AR change
            cf_inv_change = get_val("CF_CHAGE_IN_INVENTORIES", year) # Cash effect from Inv change
            # Assuming ACCT_PAYABLE_ACCRUALS_DETAILED is used as a proxy or component.
            # This part of the calculation might need refinement based on specific accounting definitions.
            # The original formula used BS_ACCT_PAYABLE for the AP component.
            # If we are using CF fields, we'd need CF_CHANGE_IN_PAYABLES.
            # Let's assume for now it's a placeholder or needs specific CF field for AP change.
            # For simplicity, if the intent is to use the BS AP value directly in the formula (which is unusual for a CF reconciliation item):
            bs_ap_val = get_val("ACCT_PAYABLE_ACCRUALS_DETAILED", year) # This is a balance, not a change.
                                                                       # This makes the formula hard to interpret in standard CF terms.
                                                                       # Reverting to the user's provided formula structure:
            derived["Increase (Decrease) in Other"][year] = chg_nwc_val - (cf_ar_change + cf_inv_change + bs_ap_val)
        else:
            derived["Increase (Decrease) in Other"][year] = 0 # Or None, if Chg NWC couldn't be calculated


    return derived

def calculate_cagr(start_value, end_value, years):
    """Calculate Compound Annual Growth Rate."""
    if not isinstance(start_value, (int, float)) or not isinstance(end_value, (int, float)):
        return 0
    if start_value == 0 or years <= 0: # Avoid division by zero or invalid year count
        return 0
    if (end_value >= 0 and start_value < 0) or (end_value < 0 and start_value > 0): # Sign change, CAGR not meaningful
        return "N/M" # Not Meaningful
    if start_value < 0 and end_value < 0: # Both negative, flip signs for calculation logic
        start_value, end_value = abs(start_value), abs(end_value)
        # Result interpretation might need care if CAGR is for negative numbers.
        # Standard CAGR implies growth from a positive base.
        # ((end / start) ^ (1/n)) - 1. If end < start (more negative), ratio < 1, CAGR negative.
        # If end > start (less negative), ratio > 1, CAGR positive.
        return ((end_value / start_value) ** (1 / years) - 1) * 100 * -1 # Re-invert if needed based on context

    return ((end_value / start_value) ** (1 / years) - 1) * 100

# Updated field map with corrected Bloomberg fields
field_map = {
    # Income Statement (IS)
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "ARDR_DEPRECIATION_AMORTIZATION", "statement": "IS"}, # Combined D&A
    "Depreciation Expense": {"source": "BDH", "field": "CF_DEPR", "statement": "IS"}, # Often taken from CF
    "Amortization Expense": {"source": "BDH", "field": "AMORT_INTANG", "statement": "IS"}, # Often taken from CF
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPERATING_INCOME", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"}, # Also used for CF start
    "EPS Basic": {"source": "BDH", "field": "IS_EPS", "statement": "IS"}, # IS_EPS for basic
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_BASIC_AVG_SHARES", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_DILUTED_AVG_SHARES", "statement": "IS"},
    # Balance Sheet (BS)
    "Cash & Cash Equivalents & ST Investments": {"source": "BDH", "field": "CASH_EQ_STI", "statement": "BS"}, # CASH_EQ_STI more common
    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS"},
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS"},
    "Prepaid Expenses and Other Current Assets": {"source": "BDH", "field": "BS_PREPAID_EXP", "statement": "BS"}, # BS_PREPAID_EXP more specific
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS"},
    "Net PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_NET_FIX_ASSET", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS"},
    "Right-of-Use Assets": {"source": "BDH", "field": "OPER_LEASE_ASSETS", "statement": "BS"}, # For operating leases under new standards
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLE", "statement": "BS"}, # Total Intangibles
    "Goodwill": {"source": "BDH", "field": "GOODWILL", "statement": "BS"},
    "Intangibles excl. Goodwill": {"source": "BDH", "field": "NET_OTHER_INTAN_ASSETS", "statement": "BS"},
    "Other Non-Current Assets": {"source": "BDH", "field": "OTH_NON_CUR_ASSETS", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSETS", "statement": "BS"},
    "Total Assets": {"source": "BDH", "field": "BS_TOT_ASSET", "statement": "BS"}, # BS_TOT_ASSET common
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS"}, # BS_ACCT_PAYABLE more specific
    "Short-Term Debt": {"source": "BDH", "field": "BS_ST_BORROW", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "SHORT_TERM_DEBT_DETAILED", "statement": "BS"}, # More detailed if needed
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "ST_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Accrued Expenses and Other Current Liabilities": {"source": "BDH", "field": "BS_ACCURED_EXP", "statement": "BS"}, # BS_ACCURED_EXP more specific
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"},
    "Long-Term Debt": {"source": "BDH", "field": "BS_LT_BORROW", "statement": "BS"},
    # Cash Flow Statement (CF) - Fields are typically changes or actual cash flows
    "Net Income (CF Start)": {"source": "BDH", "field": "NET_INCOME", "statement": "CF", "section": "Operating"}, # Re-iterating for clarity in CF section
    "D&A (CF)": {"source": "BDH", "field": "CF_DEPR_AMORT", "statement": "CF", "section": "Operating"}, # CF_DEPR_AMORT often used
    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHAGE_IN_INVENTORIES", "statement": "CF", "section": "Operating"},
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMPENSATION", "statement": "CF", "section": "Operating"},
    "Other Operating Adjustments": {"source": "BDH", "field": "CF_OTHER_OPERATING_ACT", "statement": "CF", "section": "Operating"}, # Catch-all
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    "Net Capex": {"source": "BDH", "field": "CAPITAL_EXPEND", "statement": "CF", "section": "Investing"}, # CAPITAL_EXPEND is common
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "CF_PURCHASE_OF_FIXED_PROD_ASSETS", "statement": "CF", "section": "Investing"},
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "CF_DISPOSAL_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_CASH_FOR_DIVESTURES", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_INCRE_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECRE_INVEST", "statement": "CF", "section": "Investing"},
    "Other Investing Inflows (Outflows)": {"source": "BDH", "field": "OTHER_INVESTING_ACT_DETAILED", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    "Lease Payments": {"source": "BDH", "field": "CF_LEASE_PAYMENTS", "statement": "CF", "section": "Financing"}, # Principal portion
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PROCEEDS", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PAYMENT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "CF_COMMON_STOCK_ISSUED_REPURCH", "statement": "CF", "section": "Financing"}, # Net issuance/repurchase
    "Other Financing Inflows (Outflows)": {"source": "BDH", "field": "OTHER_FIN_AND_DEC_CAP", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CASH_FLOW_FDS_FIN_ACT", "statement": "CF", "section": "Financing"}, # CASH_FLOW_FDS_FIN_ACT common
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FOREIGN_EXCHANGES", "statement": "CF", "section": "All"}, # Applies to overall CF reconciliation
    # Additional Fields (often for valuation context, can be BS or calculated)
    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"}, # CUR_MKT_CAP for latest, HISTORICAL_MARKET_CAP for history
    "Total Debt": {"source": "BDH", "field": "SHORT_AND_LONG_TERM_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "BS_PFD_EQTY_AND_HYBRID_CAP", "statement": "BS"}, # BS_PFD_EQTY_AND_HYBRID_CAP
    "Non-Controlling Interest": {"source": "BDH", "field": "BS_MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"}, # BS_MINORITY_NONCONTROLLING_INTEREST
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"}, # Can be historical
    "Total Borrowings": {"source": "BDH", "field": "TOT_BORROW", "statement": "BS"}, # TOT_BORROW
    "Total Leases": {"source": "BDH", "field": "BS_TOT_LEASE_LIAB", "statement": "BS"}, # BS_TOT_LEASE_LIAB
    "Net Debt": {"source": "BDH", "field": "NET_DEBT", "statement": "BS"},
    "Effective Tax Rate": {"source": "BDH", "field": "EFFECTIVE_TAX_RATE", "statement": "IS"}, # Usually an IS metric
    # Derived Metrics (calculated in the script)
    "Changes in Net Working Capital": {"source": "derived", "field": "Changes in Net Working Capital", "statement": "CF", "section": "Operating"}, # Often considered part of Op CF
    "DSO": {"source": "derived", "field": "DSO", "statement": "BS"}, # Or IS context (efficiency ratios)
    "DIH": {"source": "derived", "field": "DIH", "statement": "BS"},
    "DPO": {"source": "derived", "field": "DPO", "statement": "BS"},
    "Net Cash from Investments & Acquisitions": {"source": "derived", "field": "Net Cash from Investments & Acquisitions", "statement": "CF", "section": "Investing"},
    "Increase (Decrease) in Other": {"source": "derived", "field": "Increase (Decrease) in Other", "statement": "CF", "section": "Operating"}
}


# Manual cell mapping for 2014 data (column G)
# Ensure these row numbers match your Excel template EXACTLY.
field_cell_map = {
    # Income Statement (IS) - Rows 6-28 in example
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Other Operating (Income) Expenses": "G11",
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13", # Combined D&A
    "Depreciation Expense": "G14", # Specific Depreciation
    "Amortization Expense": "G15", # Specific Amortization
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24", # IS Net Income
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",

    # Balance Sheet (BS) - Rows 32-63 in example (adjust as per your template)
    "Cash & Cash Equivalents & ST Investments": "G32",
    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    "Prepaid Expenses and Other Current Assets": "G37",
    "Current Assets": "G38",
    "Net PP&E (Property, Plant and Equipment)": "G39",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    "Right-of-Use Assets": "G42",
    "Intangibles": "G43",
    "Goodwill": "G44",
    "Intangibles excl. Goodwill": "G45",
    "Other Non-Current Assets": "G46",
    "Non-Current Assets": "G47",
    "Total Assets": "G48",
    "Accounts Payable": "G49",
    "Short-Term Debt": "G50", # Often combined with current portion of LT debt
    "Short-Term Borrowings": "G51", # More specific if available
    "Current Portion of Lease Liabilities": "G52",
    "Accrued Expenses and Other Current Liabilities": "G53",
    "Current Liabilities": "G54",
    "Long-Term Debt": "G55",
    # Add other BS items as needed, e.g., Total Liabilities, Equity items, etc.
    # "Total Liabilities": "G56",
    # "Total Equity": "G60",
    # "Total Liabilities & Equity": "G63",

    # Cash Flow Statement (CF) - Rows 66-90 in example
    "Net Income (CF Start)": "G66", # Starting point for Indirect CF
    "D&A (CF)": "G67", # Depreciation & Amortization from CF
    "(Increase) Decrease in Accounts Receivable": "G68",
    "(Increase) Decrease in Inventories": "G69",
    "Stock Based Compensation": "G70",
    "Increase (Decrease) in Other": "G71", # Derived: Chg NWC - (CF AR + CF Inv + BS AP)
    "Other Operating Adjustments": "G72", # Other non-cash adjustments from BBG
    "Operating Cash Flow": "G73",
    "Net Capex": "G74",
    "Acquisition of Fixed & Intangibles": "G75",
    "Disposal of Fixed & Intangibles": "G76",
    "Acquisitions": "G77",
    "Divestitures": "G78",
    "Increase in LT Investment": "G79",
    "Decrease in LT Investment": "G80",
    "Net Cash from Investments & Acquisitions": "G81", # Derived
    "Other Investing Inflows (Outflows)": "G82",
    "Investing Cash Flow": "G83",
    "Lease Payments": "G84",
    "Debt Borrowing": "G85",
    "Debt Repayment": "G86",
    "Dividends": "G87",
    "Increase (Repurchase) of Shares": "G88",
    "Other Financing Inflows (Outflows)": "G89",
    "Financing Cash Flow": "G90",
    "Effect of Foreign Exchange": "G91",
    # "Net Change in Cash": "G92", (Calculated: OpCF + InvCF + FinCF + FX)

    # Additional Metrics / Valuation Context - Rows 94+ in example
    "Market Capitalization": "G94",
    "Total Debt": "G95", # Sum of ST and LT debt
    "Preferred Stock": "G96",
    "Non-Controlling Interest": "G97",
    "Enterprise Value": "G98",
    "Total Borrowings": "G99", # From BBG if different from Total Debt
    "Total Leases": "G100", # Total lease liabilities
    "Net Debt": "G101",
    "Effective Tax Rate": "G102",

    # Derived Ratios / Metrics (can be placed where appropriate)
    "Changes in Net Working Capital": "G103", # Derived, often part of Op CF workings
    "DSO": "G104", # Days Sales Outstanding
    "DIH": "G105", # Days Inventory Held
    "DPO": "G106", # Days Payable Outstanding
}


def filter_field_map_for_task(task_name):
    """
    Filters the global field_map for a specific task (e.g., "IS", "BS", "CF_Operating").
    Also identifies dependent fields needed for derived metrics within that task.
    """
    statement_code, cf_section = task_name.split("_") if "_" in task_name else (task_name, None)

    # Validate statement code
    allowed_statements = ["IS", "BS", "CF"]
    if statement_code not in allowed_statements:
        raise ValueError(f"Invalid statement code '{statement_code}'. Must be one of {allowed_statements}.")

    # Validate CF section if provided
    if statement_code == "CF" and cf_section and cf_section not in ["Operating", "Investing", "Financing"]:
        raise ValueError(f"Invalid CF section '{cf_section}'. Must be one of Operating, Investing, Financing, or None for all CF fields.")

    # Initial filter based on statement and CF section
    task_specific_fields = {}
    for name, config in field_map.items():
        if config["statement"] == statement_code:
            if statement_code == "CF":
                if cf_section: # Specific CF section (e.g., CF_Operating)
                    if config.get("section") == cf_section or config.get("section") == "All": # "All" applies to any CF section
                        task_specific_fields[name] = config
                else: # Entire CF statement (task "CF")
                    task_specific_fields[name] = config
            else: # IS or BS
                task_specific_fields[name] = config
    
    # Identify BDH fields required for derived metrics within this task
    # For example, if "Changes in Net Working Capital" is part of this task, we need its BDH components.
    required_bdh_for_derived = set()
    for name, config in task_specific_fields.items():
        if config["source"] == "derived":
            # Add dependencies for this derived metric
            if name == "Changes in Net Working Capital":
                required_bdh_for_derived.update(["BS_CUR_ASSET_REPORT", "BS_CUR_LIAB"])
            elif name == "DSO":
                required_bdh_for_derived.update(["BS_ACCT_NOTE_RCV", "SALES_REV_TURN"])
            elif name == "DIH":
                required_bdh_for_derived.update(["BS_INVENTORIES", "IS_COG_AND_SERVICES_SOLD"])
            elif name == "DPO":
                required_bdh_for_derived.update(["BS_ACCT_PAYABLE", "IS_COG_AND_SERVICES_SOLD"]) # Using BS_ACCT_PAYABLE
            elif name == "Net Cash from Investments & Acquisitions":
                required_bdh_for_derived.update(["CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "CF_CASH_FOR_DIVESTURES", "OTHER_INVESTING_ACT_DETAILED"])
            elif name == "Increase (Decrease) in Other": # Based on user's original derived metric dependencies
                 required_bdh_for_derived.update([
                     "BS_CUR_ASSET_REPORT", "BS_CUR_LIAB", # For Chg NWC part
                     "CF_ACCT_RCV_UNBILLED_REV", "CF_CHAGE_IN_INVENTORIES", 
                     "BS_ACCT_PAYABLE" # BS AP as per original formula for this specific derived item
                 ])


    # Add these dependent BDH fields to the task_specific_fields if they aren't already there
    # This ensures we fetch data needed for calculations even if the dependent BDH field itself isn't directly mapped for output in this task
    for bdh_field_code in required_bdh_for_derived:
        found = False
        for name, config in task_specific_fields.items():
            if config.get("field") == bdh_field_code and config.get("source") == "BDH":
                found = True
                break
        if not found:
            # Find the original definition from global field_map and add it
            for name, global_config in field_map.items():
                if global_config.get("field") == bdh_field_code and global_config.get("source") == "BDH":
                    # Add it, but mark it as a dependency (optional, or just ensure it's fetched)
                    task_specific_fields[f"__dep_{name}"] = global_config # Prefix to avoid direct output if not mapped
                    break
    return task_specific_fields


def batch_fields(fields_to_fetch, batch_size=25):
    """Split fields into batches of batch_size or fewer."""
    # fields_to_fetch should be a list of Bloomberg field codes (e.g., "SALES_REV_TURN")
    unique_fields = sorted(list(set(fields_to_fetch))) # Ensure uniqueness and consistent order
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index): # Renamed for clarity
    """Convert 1-based column index to letter (e.g., 7 -> G)."""
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    """
    Get list of cell references for a row, for num_years, starting from base_cell_ref.
    Example: base_cell_ref='G6', num_years=11 -> ['G6', 'H6', ..., 'Q6']
    """
    try:
        col_str = "".join(filter(str.isalpha, base_cell_ref))
        row_num = int("".join(filter(str.isdigit, base_cell_ref)))
        start_col_idx = openpyxl.utils.column_index_from_string(col_str)
        
        target_cells = []
        for i in range(num_years):
            target_col_letter = get_column_letter_from_index(start_col_idx + i)
            target_cells.append(f"{target_col_letter}{row_num}")
        return target_cells
    except ValueError as e:
        print(f"[ERROR] Invalid base cell reference '{base_cell_ref}': {e}")
        raise

def populate_valuation_model(template_path, output_path, ticker_symbol):
    """
    Populates the 'Inputs' sheet of the valuation model Excel file.
    Data is fetched from Bloomberg in batches and written to the single output_path file.
    The output_path file is created from the template_path at the beginning.
    """
    if not os.path.exists(template_path):
        print(f"[ERROR] Template file '{template_path}' not found.")
        raise FileNotFoundError(f"Template file {template_path} not found.")

    # --- 1. Setup Output File (Done ONCE) ---
    try:
        shutil.copy(template_path, output_path)
        print(f"[INFO] Copied template '{template_path}' to output file '{output_path}'.")
    except Exception as e:
        print(f"[ERROR] Could not copy template to output file: {e}")
        raise

    try:
        wb = openpyxl.load_workbook(output_path)
        if "Inputs" not in wb.sheetnames:
            print("[ERROR] 'Inputs' sheet not found in the workbook.")
            raise ValueError("'Inputs' sheet not found in the template file.")
        ws = wb["Inputs"]
        print("[INFO] Successfully loaded 'Inputs' sheet from the workbook.")
    except Exception as e:
        print(f"[ERROR] Could not load workbook or 'Inputs' sheet: {e}")
        raise

    # --- 2. Define Processing Scope ---
    # Tasks define the logical sections to process (e.g., Income Statement, Balance Sheet sections)
    # These correspond to how `filter_field_map_for_task` expects input.
    tasks_to_process = ["IS", "BS", "CF_Operating", "CF_Investing", "CF_Financing"]
    
    # Years for data fetching and population in Excel
    data_years = list(range(2014, 2024 + 1)) # 2014 to 2024 (11 years)
    cagr_period_years = len(data_years) -1 if len(data_years) > 1 else 0


    # --- 3. Data Fetching (All BDH data collected first) ---
    # This dictionary will store all raw data fetched from Bloomberg, keyed by BBG field code.
    # e.g., all_fetched_bdh_data['SALES_REV_TURN'] = {2014: 1000, 2015: 1100, ...}
    all_fetched_bdh_data = {}
    
    # Create a reverse map from Bloomberg field codes back to human-readable names for error reporting
    # This map is comprehensive across all fields defined in the global field_map.
    global_bberg_code_to_name_map = {
        config["field"]: name 
        for name, config in field_map.items() 
        if config["source"] == "BDH" and "field" in config
    }

    print(f"\n[PHASE] Starting data fetching for ticker: {ticker_symbol}")
    for task_name in tasks_to_process:
        print(f"\n  [TASK] Processing data collection for: {task_name}")
        
        current_task_field_configs = filter_field_map_for_task(task_name)
        
        # Extract unique Bloomberg field codes (BDH source) required for this task and its derived metrics
        bdh_fields_for_this_task = []
        for name, config in current_task_field_configs.items():
            if config["source"] == "BDH" and "field" in config:
                bdh_fields_for_this_task.append(config["field"])
        
        # Special handling for Net Income in CF: If IS is processed, Net Income for CF start should use that.
        # We still might need NET_INCOME field if IS task isn't run or if it's used elsewhere.
        # The writing part will handle copying from IS G24 to CF G66.
        # For fetching, ensure NET_INCOME is fetched if part of any task's requirements.

        if not bdh_fields_for_this_task:
            print(f"    [INFO] No Bloomberg (BDH) fields to fetch for task '{task_name}'. Skipping.")
            continue
            
        print(f"    [INFO] Identified {len(set(bdh_fields_for_this_task))} unique BDH fields for '{task_name}'.")

        # Batch these fields to respect Bloomberg's limit (e.g., 25 fields per request)
        field_batches = batch_fields(bdh_fields_for_this_task, batch_size=25)
        print(f"    [INFO] Split into {len(field_batches)} batches for '{task_name}'.")
        
        for batch_idx, current_batch_fields in enumerate(field_batches):
            print(f"      [BATCH] Processing batch {batch_idx + 1}/{len(field_batches)} for '{task_name}' with {len(current_batch_fields)} fields.")
            # print(f"        Fields: {current_batch_fields}") # Uncomment for debugging field lists

            # Establish a new Bloomberg session for each batch
            session = None # Ensure session is defined for finally block
            try:
                session = setup_bloomberg_session(ticker_symbol)
                if not session:
                    print(f"        [ERROR] Failed to start Bloomberg session for batch {batch_idx + 1}. Skipping this batch.")
                    # Decide on error handling: continue to next batch, or stop? For now, continue.
                    continue 
                
                # Fetch data for the current batch of fields
                # Pass global_bberg_code_to_name_map for better error messages on invalid fields
                batch_data_fetched = fetch_bloomberg_data(session, ticker_symbol, current_batch_fields, global_bberg_code_to_name_map, start_year=data_years[0], end_year=data_years[-1])
                
                if batch_data_fetched is None: # Indicates a session termination or critical fetch error
                    print(f"        [ERROR] Critical error fetching data for batch {batch_idx + 1}. Data might be incomplete.")
                    # Potentially add retry logic here or raise an error to stop all processing
                elif batch_data_fetched:
                    # Merge fetched data into the main collection
                    for field_code, yearly_data in batch_data_fetched.items():
                        if field_code not in all_fetched_bdh_data:
                            all_fetched_bdh_data[field_code] = {}
                        # Careful merge: prioritize new non-None data, handle N/A strings
                        for year, value in yearly_data.items():
                            if value is not None : # and value != "N/A (Invalid Field)" and value != "N/A (Error)": # Be more selective if needed
                                all_fetched_bdh_data[field_code][year] = value
                            elif year not in all_fetched_bdh_data[field_code]: # If no existing data for this year, store the None/NA
                                all_fetched_bdh_data[field_code][year] = value


                    print(f"        [SUCCESS] Fetched data for batch {batch_idx + 1}. {len(batch_data_fetched)} fields processed.")
                else:
                    print(f"        [INFO] No data returned for batch {batch_idx + 1}, or batch was empty.")

            except Exception as e:
                print(f"        [ERROR] An unexpected error occurred while processing batch {batch_idx + 1} for '{task_name}': {e}")
                # Log detailed error, consider impact on overall process
            finally:
                if session:
                    try:
                        session.stop()
                        print(f"      [BATCH] Bloomberg session stopped for batch {batch_idx + 1}.")
                    except Exception as e_stop:
                        print(f"        [WARNING] Error stopping Bloomberg session: {e_stop}")
        print(f"  [TASK] Completed data collection for: {task_name}")

    print(f"\n[PHASE] Completed all data fetching.")
    # print("[DEBUG] All fetched BDH data:", all_fetched_bdh_data) # For debugging

    # --- 4. Calculate All Derived Metrics ---
    print(f"\n[PHASE] Calculating derived metrics...")
    # This uses the `all_fetched_bdh_data` which contains everything collected.
    all_derived_data = calculate_derived_metrics(all_fetched_bdh_data, start_year=data_years[0], end_year=data_years[-1])
    print("[INFO] Derived metrics calculated.")
    # print("[DEBUG] All derived data:", all_derived_data) # For debugging

    # --- 5. Write Data to Excel Sheet ---
    print(f"\n[PHASE] Writing all data to Excel sheet '{ws.title}'...")

    # Special handling for Net Income in Cash Flow Statement:
    # It should be copied from the Income Statement's Net Income (G24 and subsequent year columns).
    # This needs to happen *before* the main loop writes CF fields, if G66 is NET_INCOME.
    is_net_income_base_cell = field_cell_map.get("Net Income") # e.g., G24
    cf_net_income_base_cell = field_cell_map.get("Net Income (CF Start)") # e.g., G66

    if is_net_income_base_cell and cf_net_income_base_cell:
        print(f"  [INFO] Copying Net Income from IS ({is_net_income_base_cell} range) to CF ({cf_net_income_base_cell} range).")
        is_ni_cells = get_target_cells_for_years(is_net_income_base_cell, len(data_years))
        cf_ni_cells = get_target_cells_for_years(cf_net_income_base_cell, len(data_years))
        
        is_ni_values_for_cagr = []
        for idx, (is_cell_ref, cf_cell_ref) in enumerate(zip(is_ni_cells, cf_ni_cells)):
            if ws[is_cell_ref].value is not None and isinstance(ws[is_cell_ref].value, (int, float)):
                val_to_copy = ws[is_cell_ref].value # Assume it's already scaled if IS was written by now
                                                   # This implies IS needs to be written first OR we use raw data
                # Let's use raw fetched data for IS Net Income to ensure consistency before scaling
                # This copy should ideally happen AFTER IS data is written OR use raw data.
                # To be safe, let's assume IS data (G24 etc.) isn't populated yet by this loop.
                # We need to get the raw 'NET_INCOME' data for the IS line.
                # This is tricky because this writing loop iterates through tasks.
                # Alternative: populate G66 using all_fetched_bdh_data['NET_INCOME'] directly.
                
                # Simplification: The main loop below will handle writing NET_INCOME to G24.
                # Then, if we encounter "Net Income (CF Start)", we copy from G24.
                # This requires "Net Income" (IS) to be processed before "Net Income (CF Start)" (CF).
                # The current `field_map` has "Net Income" for IS and "Net Income (CF Start)" for CF.
                # The `populate_valuation_model` loop iterates `tasks_to_process`.
                # If "IS" task runs and writes to G24, then later "CF_Operating" task runs,
                # it can then copy from G24 to G66.

                # This pre-copying step is better done by ensuring NET_INCOME for CF is sourced from IS data.
                # The current loop iterates tasks and then fields.
                # The "Net Income (CF Start)" entry in field_map will be handled like any other field.
                # We need a special logic for it.
                pass # This specific copy logic will be handled inside the main writing loop below.
    
    # Iterate through all defined fields in the global field_map to ensure all mapped cells are considered.
    for item_name, config in field_map.items():
        if item_name.startswith("__dep_"): # Skip internal dependency fields
            continue

        base_cell_ref = field_cell_map.get(item_name)
        if not base_cell_ref:
            # print(f"  [DEBUG] No cell mapping for '{item_name}' in field_cell_map. Skipping Excel write for this item.")
            continue

        # print(f"  [ITEM] Processing Excel write for: '{item_name}' to base cell '{base_cell_ref}'")
        target_cells_for_item = get_target_cells_for_years(base_cell_ref, len(data_years))
        yearly_values_for_cagr = []

        # Special case: "Net Income (CF Start)" should copy from IS "Net Income" cells
        if item_name == "Net Income (CF Start)" and is_net_income_base_cell:
            is_ni_source_cells = get_target_cells_for_years(is_net_income_base_cell, len(data_years))
            print(f"    [SPECIAL] Copying Net Income for CF from IS cells {is_ni_source_cells[0]}... to {target_cells_for_item[0]}...")
            for i, dest_cell_ref in enumerate(target_cells_for_item):
                source_cell_ref = is_ni_source_cells[i]
                source_value = ws[source_cell_ref].value
                if source_value is not None: # Already scaled and formatted from IS write
                    ws[dest_cell_ref] = source_value
                    ws[dest_cell_ref].number_format = ws[source_cell_ref].number_format or "#,##0.000"
                    if isinstance(source_value, (int, float)):
                         yearly_values_for_cagr.append(source_value)
                    else:
                         yearly_values_for_cagr.append(0) # Treat non-numeric as 0 for CAGR
                else:
                    ws[dest_cell_ref] = None # Or 0, or "N/A"
                    yearly_values_for_cagr.append(0)
        
        elif config["source"] == "BDH":
            bberg_field_code = config["field"]
            # print(f"    Source: BDH, Field Code: {bberg_field_code}")
            data_source_for_item = all_fetched_bdh_data.get(bberg_field_code, {})
            for i, year in enumerate(data_years):
                cell_ref = target_cells_for_item[i]
                raw_value = data_source_for_item.get(year)
                
                if isinstance(raw_value, (int, float)):
                    # Scale: Bloomberg data often in thousands or actuals. Model might expect millions.
                    # Assuming model expects millions, and BBG provides actuals for financials (divide by 1,000,000)
                    # Or if BBG provides thousands (divide by 1,000 for millions)
                    # Let's assume financial data from BBG is in actuals, and we want millions.
                    # Example: Revenue 1,234,567,890 -> display as 1,234.568 (millions)
                    # If BBG gives 1234567 (thousands), then 1234.567 (millions) is value / 1000.
                    # The original script divided by 1000. Let's stick to that.
                    scaled_value = raw_value / 1000.0 
                    ws[cell_ref] = scaled_value
                    ws[cell_ref].number_format = "#,##0.000" # 3 decimal places for millions
                    yearly_values_for_cagr.append(scaled_value)
                elif isinstance(raw_value, str) and "N/A" in raw_value: # Handle N/A strings from fetch
                    ws[cell_ref] = raw_value 
                    yearly_values_for_cagr.append(0) # Treat N/A as 0 for CAGR
                elif raw_value is None:
                    ws[cell_ref] = None # Or 0 or "N/A"
                    yearly_values_for_cagr.append(0) # Treat None as 0 for CAGR
                else: # Non-numeric, non-N/A string (e.g. some error message)
                    ws[cell_ref] = str(raw_value)
                    yearly_values_for_cagr.append(0)


        elif config["source"] == "derived":
            # print(f"    Source: Derived, Field Name: {config['field']}")
            data_source_for_item = all_derived_data.get(config["field"], {})
            for i, year in enumerate(data_years):
                cell_ref = target_cells_for_item[i]
                value = data_source_for_item.get(year)
                if isinstance(value, (int, float)):
                    ws[cell_ref] = value
                    # Formatting for derived metrics can vary. Example:
                    if "DSO" in item_name or "DIH" in item_name or "DPO" in item_name:
                        ws[cell_ref].number_format = "0.0" # Days, 1 decimal
                    else: # e.g. Changes in NWC
                        ws[cell_ref].number_format = "#,##0.000" # Financial value
                    yearly_values_for_cagr.append(value)
                elif value is None:
                    ws[cell_ref] = None
                    yearly_values_for_cagr.append(0)
                else:
                    ws[cell_ref] = str(value) # If derived metric returns non-numeric
                    yearly_values_for_cagr.append(0)

        # Calculate and write CAGR if applicable
        if yearly_values_for_cagr and cagr_period_years > 0:
            # CAGR uses the first and last values of the period.
            start_value_cagr = yearly_values_for_cagr[0]
            # Ensure we use the value for the last year in `data_years` for end_value_cagr
            # This corresponds to `yearly_values_for_cagr[len(data_years) - 1]`
            # which is `yearly_values_for_cagr[cagr_period_years]`
            end_value_cagr = yearly_values_for_cagr[cagr_period_years]


            # Determine CAGR cell (column to the right of the last year's data)
            last_year_cell_ref = target_cells_for_item[-1]
            last_year_col_str = "".join(filter(str.isalpha, last_year_cell_ref))
            row_num_str = "".join(filter(str.isdigit, last_year_cell_ref))
            cagr_col_idx = openpyxl.utils.column_index_from_string(last_year_col_str) + 1
            cagr_col_letter = get_column_letter_from_index(cagr_col_idx)
            cagr_cell_ref = f"{cagr_col_letter}{row_num_str}"
            
            cagr_value = calculate_cagr(start_value_cagr, end_value_cagr, cagr_period_years)
            if isinstance(cagr_value, str): # e.g. "N/M"
                 ws[cagr_cell_ref] = cagr_value
            else: # Numeric CAGR
                 ws[cagr_cell_ref] = cagr_value / 100.0 # Convert percentage to decimal for Excel format
                 ws[cagr_cell_ref].number_format = "0.00%" # Percentage format
            # print(f"    CAGR for '{item_name}' ({start_value_cagr} to {end_value_cagr} over {cagr_period_years} yrs) = {cagr_value}%, written to {cagr_cell_ref}")


    # --- 6. Save Final Workbook (ONCE at the end) ---
    try:
        wb.save(output_path)
        print(f"\n[SUCCESS] Valuation model populated and saved to '{output_path}'")
    except Exception as e:
        print(f"[ERROR] Failed to save the final workbook to '{output_path}': {e}")
        # Consider implications: partial data might be lost if overwrite failed.
        # Original output_path might still be the copied template if save fails here.
        raise

if __name__ == "__main__":
    print("-" * 70)
    print("Bloomberg Data to Excel Valuation Model Populator")
    print("-" * 70)
    print("This script fetches financial data from Bloomberg using blpapi,")
    print("processes it in batches, calculates derived metrics, and populates")
    print("an Excel template. The output is a single updated Excel file.")
    print("Ensure Bloomberg Terminal is running and blpapi is correctly configured.")
    print("-" * 70)

    # Configuration
    # Path to your empty Excel template file
    excel_template_path = "LIS_Valuation_Empty.xlsx" 
    # Directory for output files (optional, defaults to script directory)
    output_directory = "." 

    # Get ticker symbol from user
    ticker_input = ""
    while not ticker_input:
        ticker_input = input("Enter the Ticker Symbol (e.g., AAPL for Apple Inc.): ").strip().upper()
        if not ticker_input:
            print("[VALIDATION] Ticker symbol cannot be empty. Please try again.")
        elif not ticker_input.isalnum(): # Basic check, might need to be more lenient for some tickers
            print("[VALIDATION] Ticker symbol should ideally be alphanumeric. Please re-enter or confirm.")
            # Allow user to proceed if they confirm non-alphanumeric ticker
            confirm = input(f"Ticker '{ticker_input}' contains non-alphanumeric characters. Proceed? (y/n): ").strip().lower()
            if confirm != 'y':
                ticker_input = "" # Reset to ask again
    
    # Construct the output file path
    output_file_name = f"{ticker_input}_Valuation_Model_{datetime.now().strftime('%Y%m%d')}.xlsx"
    final_output_path = os.path.join(output_directory, output_file_name)

    print(f"\n[SETUP] Template: '{excel_template_path}'")
    print(f"[SETUP] Output will be: '{final_output_path}'")
    print(f"[SETUP] Ticker: '{ticker_input}'")

    try:
        print("\nStarting the data population process...\n")
        populate_valuation_model(
            template_path=excel_template_path,
            output_path=final_output_path,
            ticker_symbol=ticker_input
        )
        print("\nProcess completed successfully.")

    except FileNotFoundError as e_fnf:
        print(f"[CRITICAL ERROR] File not found: {e_fnf}. Please check template path.")
    except ValueError as e_val:
        print(f"[CRITICAL ERROR] Value error: {e_val}. Check configurations or data.")
    except blpapi.exception.BlpapiException as e_blp:
        print(f"[CRITICAL BLPAPI ERROR] Bloomberg API Exception: {e_blp}. Ensure Bloomberg Terminal is running and logged in.")
    except Exception as e_main:
        print(f"[UNHANDLED CRITICAL ERROR] An unexpected error occurred: {e_main}")
        import traceback
        print("\n--- Traceback ---")
        traceback.print_exc()
        print("--- End Traceback ---\n")
    finally:
        print("\nScript execution finished.")
